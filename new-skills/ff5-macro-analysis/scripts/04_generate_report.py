"""
04_generate_report.py
=====================
グレンジャー因果性検定およびランダムフォレストの分析結果を
論文用Excelレポートとしてまとめる。

出力内容:
  Sheet 1: p値マトリックス（論文 表4.3相当）
  Sheet 2: グレンジャー検定 詳細結果
  Sheet 3: ランダムフォレスト 評価サマリー
  Sheet 4: 特徴量重要度（全ファクター比較）
  Sheet 5: 特徴量重要度チャート（ヒートマップ）

使用方法:
    python 04_generate_report.py \
        --granger   results/granger_results.csv \
        --granger-matrix results/granger_results_pmatrix.csv \
        --rf        results/rf_results.csv \
        --importance results/rf_results_feature_importance.csv \
        --output    outputs/ff5_analysis_report.xlsx
"""

import argparse
import os
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# ─── 色定義 ──────────────────────────────────────────────────────────────────

COLOR_HEADER    = "1F4E79"   # 濃い青（ヘッダー背景）
COLOR_HEADER_FG = "FFFFFF"   # 白（ヘッダー文字）
COLOR_BONF_SIG  = "FF4444"   # 赤（Bonferroni補正後有意）
COLOR_PRE_SIG   = "FFC000"   # オレンジ（補正前のみ有意）
COLOR_LIGHT_BG  = "EEF3FF"   # 薄青（行の色分け）
COLOR_POS_R2    = "C6EFCE"   # 薄緑（正のOOS R²）
COLOR_NEG_R2    = "FFCCCC"   # 薄赤（負のOOS R²）


def style_header(ws, row_num: int, n_cols: int):
    """ヘッダー行のスタイルを設定する。"""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font  = Font(bold=True, color=COLOR_HEADER_FG, name="Yu Gothic UI", size=10)
        cell.fill  = PatternFill(fgColor=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_border(ws, min_row, max_row, min_col, max_col):
    """範囲にボーダーを設定する。"""
    thin = Side(style="thin", color="B0B0B0")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1, include_index: bool = True):
    """DataFrameをワークシートに書き込む。"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=include_index, header=True)):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=start_row + r_idx - 1, column=c_idx, value=value)


# ─── Sheet 1: p値マトリックス ──────────────────────────────────────────────

def build_pvalue_matrix_sheet(wb: Workbook, granger_df: pd.DataFrame, alpha: float = 0.05):
    """
    論文 表4.3「グレンジャー因果性検定のp値マトリックス」に相当するシートを作成する。
    """
    ws = wb.create_sheet("p値マトリックス（表4.3）")

    factors    = ["Mkt-RF", "SMB", "HML", "RMW", "CMA"]
    macro_vars = ["CPI_growth", "IP_growth", "TERM_SPREAD", "DEF_SPREAD",
                  "FX_return", "OIL_return", "VIX"]

    n_tests   = len(granger_df.dropna(subset=["p値"]))
    alpha_bonf = alpha / max(n_tests, 1)

    # タイトル
    ws["A1"] = "表4.3：グレンジャー因果性検定のp値マトリックス"
    ws["A1"].font = Font(bold=True, size=12, name="Yu Gothic UI")
    ws["A2"] = f"ラグ次数はBICにより選択（最大6ヶ月）。★ Bonferroni補正後も有意（p < {alpha_bonf:.4f}）、* 補正前p < 0.05"
    ws["A2"].font = Font(size=9, name="Yu Gothic UI", color="666666")

    # ヘッダー行
    start_row = 4
    ws.cell(row=start_row, column=1, value="マクロ変数 \\ ファクター")
    for c_idx, factor in enumerate(factors, 2):
        ws.cell(row=start_row, column=c_idx, value=factor)
    style_header(ws, start_row, len(factors) + 1)

    # データ行
    pivot = granger_df.pivot_table(
        index="マクロ変数", columns="ファクター", values="p値", aggfunc="first"
    )
    bonf_pivot = granger_df.pivot_table(
        index="マクロ変数", columns="ファクター", values="Bonferroni補正後有意", aggfunc="first"
    )
    pre_pivot = granger_df.pivot_table(
        index="マクロ変数", columns="ファクター", values="補正前有意（p<0.05）", aggfunc="first"
    )

    for r_idx, macro_var in enumerate(macro_vars):
        row_num = start_row + 1 + r_idx
        ws.cell(row=row_num, column=1, value=macro_var)
        ws.cell(row=row_num, column=1).font = Font(bold=True, name="Yu Gothic UI", size=10)

        for c_idx, factor in enumerate(factors, 2):
            cell = ws.cell(row=row_num, column=c_idx)
            try:
                p_val = pivot.loc[macro_var, factor]
                is_bonf = bonf_pivot.loc[macro_var, factor]
                is_pre  = pre_pivot.loc[macro_var, factor]
            except (KeyError, TypeError):
                p_val = None
                is_bonf = False
                is_pre  = False

            if pd.isna(p_val) or p_val is None:
                cell.value = "N/A"
            elif is_bonf:
                cell.value = f"{p_val:.3f} ★"
                cell.fill  = PatternFill(fgColor=COLOR_BONF_SIG, fill_type="solid")
                cell.font  = Font(bold=True, color="FFFFFF", name="Yu Gothic UI")
            elif is_pre:
                cell.value = f"{p_val:.3f} *"
                cell.fill  = PatternFill(fgColor=COLOR_PRE_SIG, fill_type="solid")
                cell.font  = Font(bold=True, name="Yu Gothic UI")
            else:
                cell.value = round(p_val, 3)
                cell.font  = Font(name="Yu Gothic UI", size=10)

            cell.alignment = Alignment(horizontal="center")

        # 交互行の色
        if r_idx % 2 == 1:
            for c in range(1, len(factors) + 2):
                if ws.cell(row=row_num, column=c).fill.fgColor.rgb == "00000000":
                    ws.cell(row=row_num, column=c).fill = PatternFill(
                        fgColor=COLOR_LIGHT_BG, fill_type="solid"
                    )

    set_border(ws, start_row, start_row + len(macro_vars), 1, len(factors) + 1)

    # 凡例
    legend_row = start_row + len(macro_vars) + 2
    ws.cell(row=legend_row, column=1, value="凡例:")
    ws.cell(row=legend_row, column=2, value="★ = Bonferroni補正後も有意")
    ws.cell(row=legend_row, column=2).fill = PatternFill(fgColor=COLOR_BONF_SIG, fill_type="solid")
    ws.cell(row=legend_row, column=2).font = Font(color="FFFFFF", name="Yu Gothic UI")
    ws.cell(row=legend_row + 1, column=2, value="* = 補正前のみ有意（参考）")
    ws.cell(row=legend_row + 1, column=2).fill = PatternFill(fgColor=COLOR_PRE_SIG, fill_type="solid")

    # 列幅
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(factors) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


# ─── Sheet 2: グレンジャー詳細結果 ─────────────────────────────────────────

def build_granger_detail_sheet(wb: Workbook, granger_df: pd.DataFrame):
    ws = wb.create_sheet("グレンジャー検定_詳細")

    ws["A1"] = "グレンジャー因果性検定 詳細結果"
    ws["A1"].font = Font(bold=True, size=12, name="Yu Gothic UI")

    cols_to_show = ["マクロ変数", "ファクター", "BIC選択ラグ", "F統計量", "p値",
                    "補正前有意（p<0.05）", "Bonferroni補正後有意", "有意度", "観測数"]
    cols_to_show = [c for c in cols_to_show if c in granger_df.columns]
    df_show = granger_df[cols_to_show].sort_values(["ファクター", "p値"])

    write_df_to_sheet(ws, df_show, start_row=3, include_index=False)
    style_header(ws, 3, len(cols_to_show))

    for col, width in zip(range(1, len(cols_to_show) + 1), [18, 10, 12, 10, 10, 16, 18, 25, 8]):
        ws.column_dimensions[get_column_letter(col)].width = width


# ─── Sheet 3: ランダムフォレスト サマリー ─────────────────────────────────

def build_rf_summary_sheet(wb: Workbook, rf_df: pd.DataFrame):
    ws = wb.create_sheet("RF_評価サマリー")

    ws["A1"] = "ランダムフォレスト評価サマリー（ウォークフォワード検証）"
    ws["A1"].font = Font(bold=True, size=12, name="Yu Gothic UI")
    ws["A2"] = "OOS R² > 0 → ヒストリカル平均より優れた予測。方向的一致率 > 0.5 → ランダムより良い予測。"
    ws["A2"].font = Font(size=9, color="666666", name="Yu Gothic UI")

    write_df_to_sheet(ws, rf_df, start_row=4, include_index=False)
    style_header(ws, 4, len(rf_df.columns))

    # OOS R² のセルに色付け
    oos_col_all = rf_df.columns.get_loc("OOS_R2（全期間）") + 1 if "OOS_R2（全期間）" in rf_df.columns else None
    oos_col_vix = rf_df.columns.get_loc("OOS_R2（VIX含む）") + 1 if "OOS_R2（VIX含む）" in rf_df.columns else None

    for r_idx, (_, row) in enumerate(rf_df.iterrows(), 5):
        for col_idx in [oos_col_all, oos_col_vix]:
            if col_idx is None:
                continue
            col_name = rf_df.columns[col_idx - 1]
            val = row.get(col_name)
            if pd.notna(val):
                cell = ws.cell(row=r_idx, column=col_idx)
                if float(val) > 0:
                    cell.fill = PatternFill(fgColor=COLOR_POS_R2, fill_type="solid")
                else:
                    cell.fill = PatternFill(fgColor=COLOR_NEG_R2, fill_type="solid")

    for col in range(1, len(rf_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


# ─── Sheet 4: 特徴量重要度 ─────────────────────────────────────────────────

def build_importance_sheet(wb: Workbook, imp_df: pd.DataFrame):
    ws = wb.create_sheet("特徴量重要度_比較")

    ws["A1"] = "特徴量重要度（全ファクター比較）- ランダムフォレスト"
    ws["A1"].font = Font(bold=True, size=12, name="Yu Gothic UI")
    ws["A2"] = "各ファクターのTop-10特徴量。長ラグの変数が上位 → 遅行的反応の示唆。"
    ws["A2"].font = Font(size=9, color="666666", name="Yu Gothic UI")

    # ファクター別Top-10のピボット
    top10_list = []
    factors = ["Mkt-RF", "SMB", "HML", "RMW", "CMA"]
    for factor in factors:
        sub = imp_df[imp_df["ファクター"] == factor].head(10)
        for _, row in sub.iterrows():
            top10_list.append({
                "ファクター":   factor,
                "順位":         row.get("順位", ""),
                "特徴量":       row.get("特徴量", ""),
                "マクロ変数":   row.get("マクロ変数", ""),
                "ラグ（ヶ月）": row.get("ラグ（ヶ月）", ""),
                "重要度":       round(row.get("重要度", 0), 6),
            })

    if top10_list:
        df_show = pd.DataFrame(top10_list)
        write_df_to_sheet(ws, df_show, start_row=4, include_index=False)
        style_header(ws, 4, len(df_show.columns))

        for col, width in zip(range(1, len(df_show.columns) + 1), [10, 6, 25, 18, 14, 10]):
            ws.column_dimensions[get_column_letter(col)].width = width

    # 平均ラグ（遅行度指標）のサマリー
    summary_start = 4 + len(top10_list) + 3
    ws.cell(row=summary_start, column=1, value="ファクター別 Top-10特徴量の平均ラグ（遅行度指標）")
    ws.cell(row=summary_start, column=1).font = Font(bold=True, name="Yu Gothic UI")

    for r_idx, factor in enumerate(factors, summary_start + 1):
        sub = imp_df[imp_df["ファクター"] == factor].head(10)
        avg_lag = sub["ラグ（ヶ月）"].mean() if len(sub) > 0 else np.nan
        ws.cell(row=r_idx, column=1, value=factor)
        ws.cell(row=r_idx, column=2, value=round(avg_lag, 1) if pd.notna(avg_lag) else "N/A")
        ws.cell(row=r_idx, column=3, value="ヶ月")


# ─── メイン処理 ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Excelレポート生成")
    parser.add_argument("--granger",        type=str, default="results/granger_results.csv")
    parser.add_argument("--granger-matrix", type=str, default="results/granger_results_pmatrix.csv")
    parser.add_argument("--rf",             type=str, default="results/rf_results.csv")
    parser.add_argument("--importance",     type=str, default="results/rf_results_feature_importance.csv")
    parser.add_argument("--output",         type=str, default="outputs/ff5_analysis_report.xlsx")
    parser.add_argument("--alpha",          type=float, default=0.05)
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)

    print("Excelレポートを生成中...")

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # グレンジャー結果
    if os.path.exists(args.granger):
        granger_df = pd.read_csv(args.granger, encoding="utf-8-sig", index_col=False)
        print(f"  グレンジャー結果読み込み: {len(granger_df)}行")
        build_pvalue_matrix_sheet(wb, granger_df, alpha=args.alpha)
        build_granger_detail_sheet(wb, granger_df)
    else:
        print(f"  ⚠️ グレンジャー結果が見つかりません: {args.granger}")

    # ランダムフォレスト結果
    if os.path.exists(args.rf):
        rf_df = pd.read_csv(args.rf, encoding="utf-8-sig", index_col=False)
        print(f"  RF結果読み込み: {len(rf_df)}行")
        build_rf_summary_sheet(wb, rf_df)
    else:
        print(f"  ⚠️ RF結果が見つかりません: {args.rf}")

    # 特徴量重要度
    if os.path.exists(args.importance):
        imp_df = pd.read_csv(args.importance, encoding="utf-8-sig", index_col=False)
        print(f"  特徴量重要度読み込み: {len(imp_df)}行")
        build_importance_sheet(wb, imp_df)
    else:
        print(f"  ⚠️ 特徴量重要度が見つかりません: {args.importance}")

    # 保存
    wb.save(args.output)
    print(f"\n✓ Excelレポートを保存しました: {args.output}")


if __name__ == "__main__":
    main()
