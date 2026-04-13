#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DCFバリュエーション計算スクリプト
割引キャッシュフロー法を用いた企業価値評価を自動化します。

使用方法:
    python dcf_calculator.py --input <入力ファイル> --output <出力ファイル>

入力ファイル形式:
    CSV または Excel形式で以下の列を含める必要があります:
    - Year: 年度
    - Revenue: 売上
    - Operating_Costs: 営業費用
    - D_and_A: 減価償却費
    - CapEx: 資本支出
    - Tax_Rate: 税率
    - Market_Cap: 株式時価総額
    - Debt_Value: 負債時価総額
    - Cost_of_Equity: 株式コスト
    - Cost_of_Debt: 負債コスト
"""

import argparse
import sys
import logging
from pathlib import Path
from typing import Tuple, Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ロギング設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DCFCalculator:
    """DCF(割引キャッシュフロー法)計算エンジン"""

    def __init__(
        self,
        revenue: np.ndarray,
        operating_costs: np.ndarray,
        d_and_a: np.ndarray,
        capex: np.ndarray,
        tax_rate: float,
        market_cap: float,
        debt_value: float,
        cost_of_equity: float,
        cost_of_debt: float,
        terminal_growth_rate: float = 0.025,
        forecast_years: int = 5
    ):
        """
        初期化メソッド

        Args:
            revenue: 売上配列
            operating_costs: 営業費用配列
            d_and_a: 減価償却費配列
            capex: 資本支出配列
            tax_rate: 実効税率
            market_cap: 株式時価総額
            debt_value: 負債時価総額
            cost_of_equity: 株式のコスト
            cost_of_debt: 負債のコスト
            terminal_growth_rate: 永続成長率（デフォルト: 2.5%）
            forecast_years: 予測期間（デフォルト: 5年）
        """
        self.revenue = np.asarray(revenue, dtype=float)
        self.operating_costs = np.asarray(operating_costs, dtype=float)
        self.d_and_a = np.asarray(d_and_a, dtype=float)
        self.capex = np.asarray(capex, dtype=float)
        self.tax_rate = float(tax_rate)
        self.market_cap = float(market_cap)
        self.debt_value = float(debt_value)
        self.cost_of_equity = float(cost_of_equity)
        self.cost_of_debt = float(cost_of_debt)
        self.terminal_growth_rate = float(terminal_growth_rate)
        self.forecast_years = int(forecast_years)

        # 検証
        self._validate_inputs()

        # WACC計算
        self.wacc = self._calculate_wacc()

        # FCF計算
        self.fcf_projections = self._calculate_fcf()

        # ターミナルバリュー計算
        self.terminal_value = self._calculate_terminal_value()

        # 企業価値計算
        self.enterprise_value = self._calculate_enterprise_value()

        # 感度分析用データ
        self.sensitivity_analysis = self._calculate_sensitivity_analysis()

        logger.info("DCF計算が正常に完了しました")

    def _validate_inputs(self) -> None:
        """入力値の妥当性を検証"""
        if self.tax_rate < 0 or self.tax_rate > 1:
            raise ValueError(f"税率は0から1の範囲である必要があります: {self.tax_rate}")

        if self.cost_of_equity < 0:
            raise ValueError(f"株式コストは非負である必要があります: {self.cost_of_equity}")

        if self.cost_of_debt < 0:
            raise ValueError(f"負債コストは非負である必要があります: {self.cost_of_debt}")

        if self.market_cap <= 0 or self.debt_value < 0:
            raise ValueError(f"時価総額と負債額が不正です。E={self.market_cap}, D={self.debt_value}")

        if self.wacc >= self.terminal_growth_rate:
            logger.warning(
                f"警告: WACC({self.wacc:.4f}) >= 永続成長率({self.terminal_growth_rate:.4f})"
            )

        logger.info("入力値の検証が完了しました")

    def _calculate_wacc(self) -> float:
        """
        加重平均資本コスト(WACC)を計算

        WACC = (E/(E+D)) × Re + (D/(E+D)) × Rd × (1 - Tc)
        """
        total_value = self.market_cap + self.debt_value
        equity_weight = self.market_cap / total_value
        debt_weight = self.debt_value / total_value

        wacc = (
            equity_weight * self.cost_of_equity +
            debt_weight * self.cost_of_debt * (1 - self.tax_rate)
        )

        logger.info(f"WACC計算完了: {wacc:.4f} ({wacc*100:.2f}%)")
        return wacc

    def _calculate_fcf(self) -> pd.DataFrame:
        """
        フリーキャッシュフロー(FCF)を計算

        営業利益 = 売上 - 営業費用
        NOPAT = 営業利益 × (1 - 税率)
        FCF = NOPAT + D&A - CapEx - ΔWC
        """
        # 基年度のデータを使用
        if len(self.revenue) == 0:
            raise ValueError("売上データが不足しています")

        # 過去データから成長率を推定
        if len(self.revenue) >= 2:
            growth_rate = (self.revenue[-1] / self.revenue[-2]) - 1
        else:
            growth_rate = 0.05  # デフォルト成長率

        base_revenue = self.revenue[-1]
        base_operating_margin = 1 - (self.operating_costs[-1] / self.revenue[-1])

        # 予測年度のデータを初期化
        forecast_years_data = []

        for year in range(1, self.forecast_years + 1):
            # 売上予測（成長率を適用）
            projected_revenue = base_revenue * ((1 + growth_rate) ** year)

            # 営業利益予測
            operating_profit = projected_revenue * base_operating_margin

            # NOPAT（税後営業利益）
            nopat = operating_profit * (1 - self.tax_rate)

            # D&Aの推定（売上比）
            da_ratio = self.d_and_a[-1] / self.revenue[-1] if len(self.d_and_a) > 0 else 0.05
            projected_da = projected_revenue * da_ratio

            # CapExの推定（売上比）
            capex_ratio = self.capex[-1] / self.revenue[-1] if len(self.capex) > 0 else 0.08
            projected_capex = projected_revenue * capex_ratio

            # ΔWCの推定（売上比）
            dwc_ratio = 0.02  # 売上の2%
            projected_dwc = projected_revenue * dwc_ratio

            # FCF計算
            fcf = nopat + projected_da - projected_capex - projected_dwc

            forecast_years_data.append({
                '年度': year,
                '売上': projected_revenue,
                '営業利益': operating_profit,
                'NOPAT': nopat,
                'D&A': projected_da,
                'CapEx': projected_capex,
                'ΔWC': projected_dwc,
                'FCF': fcf,
                '割引率': 1 / ((1 + self.wacc) ** year),
                '現在価値': fcf / ((1 + self.wacc) ** year)
            })

        fcf_df = pd.DataFrame(forecast_years_data)
        logger.info(f"{self.forecast_years}年間のFCF予測が完了しました")
        return fcf_df

    def _calculate_terminal_value(self) -> float:
        """
        ターミナルバリューを計算（ゴードン成長モデル）

        TV = FCFn × (1 + g) / (WACC - g)
        """
        if self.wacc <= self.terminal_growth_rate:
            logger.warning(
                f"警告: WACC <= 永続成長率。WACC={self.wacc:.4f}, g={self.terminal_growth_rate:.4f}"
            )
            # 計算を続ける（警告のみ）

        final_fcf = self.fcf_projections['FCF'].iloc[-1]

        tv = final_fcf * (1 + self.terminal_growth_rate) / (
            self.wacc - self.terminal_growth_rate
        )

        logger.info(f"ターミナルバリュー計算完了: {tv:,.0f}")
        return tv

    def _calculate_enterprise_value(self) -> Dict[str, float]:
        """
        企業価値を計算

        EV = Σ(FCFt / (1 + WACC)^t) + TV / (1 + WACC)^n
        """
        # FCFの現在価値合計
        pv_fcf = self.fcf_projections['現在価値'].sum()

        # ターミナルバリューの現在価値
        tv_discount = 1 / ((1 + self.wacc) ** self.forecast_years)
        pv_tv = self.terminal_value * tv_discount

        # 企業価値
        enterprise_value = pv_fcf + pv_tv

        logger.info(f"企業価値計算完了: {enterprise_value:,.0f}")

        return {
            'PV_FCF': pv_fcf,
            'PV_TerminalValue': pv_tv,
            'Enterprise_Value': enterprise_value,
            'Equity_Value': enterprise_value - self.debt_value,
        }

    def _calculate_sensitivity_analysis(self) -> pd.DataFrame:
        """
        感度分析（WACC vs 永続成長率のマトリックス）を計算
        """
        # WACC範囲: 6% - 12%
        wacc_range = np.arange(0.06, 0.13, 0.01)

        # 成長率範囲: 0% - 5%
        growth_range = np.arange(0.00, 0.06, 0.005)

        sensitivity_matrix = []

        for wacc in wacc_range:
            row = []
            for growth in growth_range:
                if wacc <= growth:
                    # 無効なシナリオ
                    row.append(np.nan)
                else:
                    # 企業価値を計算
                    final_fcf = self.fcf_projections['FCF'].iloc[-1]
                    tv = final_fcf * (1 + growth) / (wacc - growth)

                    pv_fcf = self.fcf_projections['現在価値'].sum() * (wacc / self.wacc)
                    tv_discount = 1 / ((1 + wacc) ** self.forecast_years)
                    pv_tv = tv * tv_discount

                    ev = pv_fcf + pv_tv
                    row.append(ev)

            sensitivity_matrix.append(row)

        sensitivity_df = pd.DataFrame(
            sensitivity_matrix,
            index=[f'{w:.1%}' for w in wacc_range],
            columns=[f'{g:.1%}' for g in growth_range]
        )

        logger.info("感度分析が完了しました")
        return sensitivity_df


def load_input_data(file_path: str) -> Tuple[
    np.ndarray, np.ndarray, np.ndarray, np.ndarray, float, float, float, float, float
]:
    """
    入力ファイルからデータを読み込む

    Args:
        file_path: CSV または Excel ファイルパス

    Returns:
        (revenue, operating_costs, d_and_a, capex, tax_rate, market_cap,
         debt_value, cost_of_equity, cost_of_debt)
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

    # ファイル形式に応じて読み込み
    if file_path.suffix.lower() == '.csv':
        df = pd.read_csv(file_path)
    elif file_path.suffix.lower() in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path)
    else:
        raise ValueError(f"サポートされていないファイル形式です: {file_path.suffix}")

    logger.info(f"入力ファイルを読み込みました: {file_path}")

    # 必須列の確認
    required_columns = [
        'Revenue', 'Operating_Costs', 'D_and_A', 'CapEx', 'Tax_Rate',
        'Market_Cap', 'Debt_Value', 'Cost_of_Equity', 'Cost_of_Debt'
    ]

    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"必須列が不足しています: {missing_columns}")

    # データ抽出
    revenue = df['Revenue'].values
    operating_costs = df['Operating_Costs'].values
    d_and_a = df['D_and_A'].values
    capex = df['CapEx'].values

    # スカラー値
    tax_rate = df['Tax_Rate'].iloc[-1]
    market_cap = df['Market_Cap'].iloc[-1]
    debt_value = df['Debt_Value'].iloc[-1]
    cost_of_equity = df['Cost_of_Equity'].iloc[-1]
    cost_of_debt = df['Cost_of_Debt'].iloc[-1]

    logger.info(f"入力データの抽出が完了しました（{len(revenue)}年分）")

    return revenue, operating_costs, d_and_a, capex, tax_rate, market_cap, debt_value, cost_of_equity, cost_of_debt


def create_excel_report(
    output_path: str,
    calculator: DCFCalculator,
    ev_dict: Dict[str, float]
) -> None:
    """
    Excelレポートを作成して保存

    Args:
        output_path: 出力ファイルパス
        calculator: DCFCalculatorインスタンス
        ev_dict: 企業価値計算結果の辞書
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Workbook作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    number_format = '#,##0.00'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. サマリーシート
    ws_summary = wb.create_sheet('サマリー', 0)

    ws_summary['A1'] = 'DCFバリュエーション分析 - サマリー'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')

    summary_data = [
        ['項目', '値'],
        ['加重平均資本コスト(WACC)', f'{calculator.wacc:.4f}'],
        ['永続成長率', f'{calculator.terminal_growth_rate:.4f}'],
        ['予測期間(年)', calculator.forecast_years],
        [''],
        ['FCFの現在価値', f'{ev_dict["PV_FCF"]:,.0f}'],
        ['ターミナルバリューの現在価値', f'{ev_dict["PV_TerminalValue"]:,.0f}'],
        ['企業価値(EV)', f'{ev_dict["Enterprise_Value"]:,.0f}'],
        ['(-) 負債', f'{calculator.debt_value:,.0f}'],
        ['(=) 株主価値', f'{ev_dict["Equity_Value"]:,.0f}'],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx > 1 and row_idx > 1 and row_data[0] != '':
                cell.number_format = number_format
            cell.border = border

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # 2. FCF予測シート
    ws_fcf = wb.create_sheet('FCF予測', 1)

    # ヘッダー
    headers = ['年度', '売上', '営業利益', 'NOPAT', 'D&A', 'CapEx', 'ΔWC', 'FCF', '割引率', '現在価値']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_fcf.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    # データ行
    for row_idx, (_, row_data) in enumerate(calculator.fcf_projections.iterrows(), 2):
        for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
            cell = ws_fcf.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_name != '年度':
                cell.number_format = number_format
            cell.alignment = Alignment(horizontal='right', vertical='center')

    # 列幅設定
    ws_fcf.column_dimensions['A'].width = 8
    for col_idx in range(2, len(headers) + 1):
        ws_fcf.column_dimensions[get_column_letter(col_idx)].width = 15

    # 3. WACC計算シート
    ws_wacc = wb.create_sheet('WACC計算', 2)

    total_value = calculator.market_cap + calculator.debt_value
    equity_weight = calculator.market_cap / total_value
    debt_weight = calculator.debt_value / total_value

    wacc_data = [
        ['項目', '値', '説明'],
        ['株式時価総額(E)', f'{calculator.market_cap:,.0f}', ''],
        ['負債時価総額(D)', f'{calculator.debt_value:,.0f}', ''],
        ['企業価値(E+D)', f'{total_value:,.0f}', ''],
        [''],
        ['株式ウェイト(E/(E+D))', f'{equity_weight:.4f}', ''],
        ['負債ウェイト(D/(E+D))', f'{debt_weight:.4f}', ''],
        [''],
        ['株式のコスト(Re)', f'{calculator.cost_of_equity:.4f}', ''],
        ['負債のコスト(Rd)', f'{calculator.cost_of_debt:.4f}', ''],
        ['実効税率(Tc)', f'{calculator.tax_rate:.4f}', ''],
        [''],
        ['WACC計算式', '', 'WACC = E/(E+D)×Re + D/(E+D)×Rd×(1-Tc)'],
        ['WACC', f'{calculator.wacc:.4f}', ''],
    ]

    for row_idx, row_data in enumerate(wacc_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_wacc.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 2 and row_data[0] != '' and row_data[0] != '計算式':
                cell.number_format = number_format
            cell.border = border

    ws_wacc.column_dimensions['A'].width = 25
    ws_wacc.column_dimensions['B'].width = 15
    ws_wacc.column_dimensions['C'].width = 50

    # 4. 感度分析シート
    ws_sensitivity = wb.create_sheet('感度分析', 3)

    ws_sensitivity['A1'] = '企業価値感度分析 (WACC vs 永続成長率)'
    ws_sensitivity['A1'].font = Font(bold=True, size=12)
    ws_sensitivity.merge_cells('A1:G1')

    # ヘッダー行
    ws_sensitivity['A3'] = 'WACC / 成長率'
    ws_sensitivity['A3'].fill = header_fill
    ws_sensitivity['A3'].font = header_font
    ws_sensitivity['A3'].border = border

    for col_idx, col_name in enumerate(calculator.sensitivity_analysis.columns, 2):
        cell = ws_sensitivity.cell(row=3, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    # 行ヘッダーとデータ
    for row_idx, (index, row_data) in enumerate(
        calculator.sensitivity_analysis.iterrows(), 4
    ):
        # 行ヘッダー
        cell = ws_sensitivity.cell(row=row_idx, column=1, value=index)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center_align

        # データ
        for col_idx, value in enumerate(row_data, 2):
            cell = ws_sensitivity.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = '#,##0'
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')

    ws_sensitivity.column_dimensions['A'].width = 15
    for col_idx in range(2, len(calculator.sensitivity_analysis.columns) + 2):
        ws_sensitivity.column_dimensions[get_column_letter(col_idx)].width = 12

    # ファイル保存
    wb.save(output_path)
    logger.info(f"Excelレポートを保存しました: {output_path}")


def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(
        description='DCF(割引キャッシュフロー法)バリュエーション計算スクリプト'
    )
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='入力ファイル（CSV または Excel形式）のパス'
    )
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='出力Excelファイルのパス'
    )
    parser.add_argument(
        '--growth-rate', '-g',
        type=float,
        default=0.025,
        help='永続成長率（デフォルト: 0.025 = 2.5%）'
    )
    parser.add_argument(
        '--forecast-years', '-f',
        type=int,
        default=5,
        help='予測期間（年数、デフォルト: 5）'
    )

    args = parser.parse_args()

    try:
        # 入力データ読み込み
        revenue, operating_costs, d_and_a, capex, tax_rate, market_cap, debt_value, cost_of_equity, cost_of_debt = load_input_data(args.input)

        # DCF計算実行
        calculator = DCFCalculator(
            revenue=revenue,
            operating_costs=operating_costs,
            d_and_a=d_and_a,
            capex=capex,
            tax_rate=tax_rate,
            market_cap=market_cap,
            debt_value=debt_value,
            cost_of_equity=cost_of_equity,
            cost_of_debt=cost_of_debt,
            terminal_growth_rate=args.growth_rate,
            forecast_years=args.forecast_years
        )

        # 企業価値計算
        ev_dict = calculator.enterprise_value

        # Excelレポート作成
        create_excel_report(args.output, calculator, ev_dict)

        # 結果表示
        print("\n" + "="*60)
        print("DCFバリュエーション分析 - 計算結果")
        print("="*60)
        print(f"加重平均資本コスト(WACC): {calculator.wacc:>15.4f} ({calculator.wacc*100:.2f}%)")
        print(f"永続成長率: {calculator.terminal_growth_rate:>30.4f} ({calculator.terminal_growth_rate*100:.2f}%)")
        print(f"ターミナルバリュー: {calculator.terminal_value:>25,.0f}")
        print("-" * 60)
        print(f"FCFの現在価値: {ev_dict['PV_FCF']:>30,.0f}")
        print(f"ターミナルバリューの現在価値: {ev_dict['PV_TerminalValue']:>15,.0f}")
        print(f"企業価値(EV): {ev_dict['Enterprise_Value']:>35,.0f}")
        print(f"(-) 負債: {calculator.debt_value:>46,.0f}")
        print(f"(=) 株主価値: {ev_dict['Equity_Value']:>35,.0f}")
        print("=" * 60)
        print(f"\n出力ファイル: {args.output}")
        print("レポート生成が完了しました！\n")

        return 0

    except FileNotFoundError as e:
        logger.error(f"ファイルエラー: {e}")
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    except ValueError as e:
        logger.error(f"値のエラー: {e}")
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    except Exception as e:
        logger.error(f"予期しないエラーが発生しました: {e}")
        print(f"予期しないエラー: {e}", file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
