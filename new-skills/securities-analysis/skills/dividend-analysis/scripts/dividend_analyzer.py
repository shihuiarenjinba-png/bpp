#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配当分析スクリプト (Dividend Analysis Script)

このスクリプトは配当データを分析し、包括的なExcelレポートを生成します。
利回り、成長率、DGM評価、再投資シミュレーション、持続可能性スコアを計算します。

使用方法:
    python dividend_analyzer.py <入力CSVパス> <出力Excelパス>

例:
    python dividend_analyzer.py ./data/dividend_data.csv ./reports/dividend_analysis.xlsx
"""

import sys
import argparse
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import rcParams

# 日本語フォント設定
rcParams['font.sans-serif'] = ['DejaVu Sans']
rcParams['axes.unicode_minus'] = False


class DividendAnalyzer:
    """配当分析メインクラス"""

    def __init__(self, csv_path: str):
        """
        初期化とCSVデータの読み込み

        Args:
            csv_path: 入力CSVファイルのパス
        """
        self.csv_path = Path(csv_path)
        self.output_excel = None
        self.df = None
        self.companies = None
        self.analyses = {}

        self._load_data()

    def _load_data(self):
        """CSVデータを読み込む"""
        try:
            self.df = pd.read_csv(self.csv_path, encoding='utf-8')

            # 列名の検証
            required_cols = ['銘柄', '年度', '配当金額', 'EPS', '株価']
            if not all(col in self.df.columns for col in required_cols):
                raise ValueError(f"必須列が不足しています。必要な列: {required_cols}")

            # データ型の変換
            self.df['年度'] = self.df['年度'].astype(int)
            self.df['配当金額'] = pd.to_numeric(self.df['配当金額'], errors='coerce')
            self.df['EPS'] = pd.to_numeric(self.df['EPS'], errors='coerce')
            self.df['株価'] = pd.to_numeric(self.df['株価'], errors='coerce')

            # NaN値を削除
            self.df = self.df.dropna()

            # 企業リストを取得
            self.companies = sorted(self.df['銘柄'].unique())

            print(f"データ読み込み成功: {len(self.companies)}社, {len(self.df)}行")

        except FileNotFoundError:
            print(f"エラー: ファイル '{self.csv_path}' が見つかりません")
            sys.exit(1)
        except Exception as e:
            print(f"エラー: データ読み込み失敗 - {str(e)}")
            sys.exit(1)

    def analyze(self):
        """全ての分析を実行"""
        for company in self.companies:
            company_df = self.df[self.df['銘柄'] == company].sort_values('年度')
            self.analyses[company] = self._analyze_company(company_df)

    def _analyze_company(self, company_df: pd.DataFrame) -> dict:
        """
        企業ごとの分析を実行

        Args:
            company_df: 企業のデータフレーム

        Returns:
            分析結果の辞書
        """
        analysis = {
            'data': company_df.copy(),
            'dividend_yield': [],
            'payout_ratio': [],
            'dividend_growth': [],
            'years_of_increase': 0,
            'cagr': 0.0,
            'ddm_valuation': {},
            'drip_simulation': {},
            'sustainability_score': 0.0
        }

        # 配当利回りと配当性向を計算
        company_df['配当利回り'] = (company_df['配当金額'] / company_df['株価']) * 100
        company_df['配当性向'] = (company_df['配当金額'] / company_df['EPS']) * 100

        analysis['dividend_yield'] = company_df['配当利回り'].tolist()
        analysis['payout_ratio'] = company_df['配当性向'].tolist()

        # 配当成長率を計算
        dividends = company_df['配当金額'].values
        analysis['dividend_growth'] = self._calculate_growth_rates(dividends)

        # 連続増配年数を計算
        analysis['years_of_increase'] = self._count_consecutive_increases(dividends)

        # CAGR（複合年間成長率）を計算
        if len(dividends) > 1:
            years = len(dividends) - 1
            analysis['cagr'] = ((dividends[-1] / dividends[0]) ** (1 / years) - 1) * 100

        # DGM（ゴードン成長モデル）評価
        analysis['ddm_valuation'] = self._calculate_ddm(
            company_df.iloc[-1],
            dividends,
            analysis['cagr']
        )

        # DRIP（配当再投資）シミュレーション
        analysis['drip_simulation'] = self._simulate_drip(
            company_df.iloc[-1]['配当金額'],
            company_df.iloc[-1]['株価'],
            analysis['cagr']
        )

        # 持続可能性スコア
        analysis['sustainability_score'] = self._calculate_sustainability_score(
            company_df['配当性向'].values,
            company_df['EPS'].values
        )

        return analysis

    def _calculate_growth_rates(self, dividends: np.ndarray) -> list:
        """配当成長率を計算"""
        growth_rates = [0.0]  # 最初の年は成長率なし
        for i in range(1, len(dividends)):
            if dividends[i-1] > 0:
                growth = ((dividends[i] - dividends[i-1]) / dividends[i-1]) * 100
                growth_rates.append(growth)
            else:
                growth_rates.append(0.0)
        return growth_rates

    def _count_consecutive_increases(self, dividends: np.ndarray) -> int:
        """連続増配年数を計算"""
        count = 0
        for i in range(len(dividends) - 1, 0, -1):
            if dividends[i] > dividends[i-1]:
                count += 1
            else:
                break
        return count

    def _calculate_ddm(self, latest_row, dividends: np.ndarray, cagr: float) -> dict:
        """
        ゴードン成長モデル（Gordon Growth Model）で理論株価を計算

        理論株価 = D₁ / (r - g)
        D₁: 次年度予想配当
        r: 要求利回り
        g: 永続成長率
        """
        result = {}

        # 次年度配当予想
        latest_dividend = latest_row['配当金額']
        d1 = latest_dividend * (1 + cagr / 100)

        # 異なる要求利回りでシミュレーション
        required_returns = [5.0, 7.0, 10.0]
        growth_rate = min(cagr / 100, 0.03)  # 永続成長率は3%を上限に

        for req_return in required_returns:
            req_return_decimal = req_return / 100
            if req_return_decimal > growth_rate:
                theoretical_price = d1 / (req_return_decimal - growth_rate)
                actual_price = latest_row['株価']
                upside = ((theoretical_price - actual_price) / actual_price) * 100

                result[f'要求利回り{int(req_return)}%'] = {
                    '理論株価': round(theoretical_price, 2),
                    '現在株価': round(actual_price, 2),
                    'アップサイド': round(upside, 2)
                }

        return result

    def _simulate_drip(self, latest_dividend: float, latest_price: float, cagr: float) -> dict:
        """
        DRIP（配当再投資）シミュレーション
        10年・20年・30年の複利成長を計算
        """
        result = {}

        initial_shares = 100  # 初期100株を仮定
        initial_value = initial_shares * latest_price
        dividend_growth_rate = cagr / 100

        for years in [10, 20, 30]:
            current_shares = initial_shares
            accumulated_value = initial_value

            for year in range(1, years + 1):
                # 年間配当を計算
                annual_dividend = latest_dividend * (1 + dividend_growth_rate) ** year
                total_dividend_income = current_shares * annual_dividend

                # 配当で新しい株を購入（株価も成長すると仮定）
                stock_price_at_year = latest_price * (1 + dividend_growth_rate) ** year
                new_shares = total_dividend_income / stock_price_at_year
                current_shares += new_shares

                accumulated_value = current_shares * stock_price_at_year

            result[f'{years}年後'] = {
                '株数': round(current_shares, 2),
                '評価額': round(accumulated_value, 2),
                '初期投資比': round(accumulated_value / initial_value, 2)
            }

        return result

    def _calculate_sustainability_score(self, payout_ratios: np.ndarray, eps_values: np.ndarray) -> float:
        """
        配当の持続可能性スコアを計算（0-100）

        要因：
        - 配当性向（平均が低いほど安全）
        - EPS安定性（変動が低いほど安全）
        """
        # 配当性向スコア（50%以下で加点、70%以上で減点）
        avg_payout = np.mean(payout_ratios)
        if avg_payout <= 30:
            payout_score = 100
        elif avg_payout <= 50:
            payout_score = 80 - (avg_payout - 30) * 1
        elif avg_payout <= 70:
            payout_score = 60 - (avg_payout - 50) * 1
        else:
            payout_score = max(0, 40 - (avg_payout - 70) * 1)

        # EPS安定性スコア
        eps_volatility = np.std(eps_values) / np.mean(eps_values) if np.mean(eps_values) > 0 else 1
        if eps_volatility < 0.1:
            stability_score = 100
        elif eps_volatility < 0.2:
            stability_score = 90 - (eps_volatility - 0.1) * 100
        elif eps_volatility < 0.5:
            stability_score = 70 - (eps_volatility - 0.2) * 100
        else:
            stability_score = max(0, 40 - (eps_volatility - 0.5) * 20)

        # 総合スコア（加重平均）
        total_score = payout_score * 0.6 + stability_score * 0.4
        return round(total_score, 1)

    def generate_excel_report(self, output_path: str):
        """Excelレポートを生成"""
        self.output_excel = Path(output_path)

        with pd.ExcelWriter(self.output_excel, engine='openpyxl') as writer:
            # シート1: サマリー
            self._create_summary_sheet(writer)

            # シート2: 配当履歴分析
            self._create_history_sheet(writer)

            # シート3: DGM評価
            self._create_dgm_sheet(writer)

            # シート4: 再投資シミュレーション
            self._create_drip_sheet(writer)

            # シート5: 持続可能性
            self._create_sustainability_sheet(writer)

        # スタイルを適用
        self._apply_styles()

        # グラフを追加
        self._add_charts()

        print(f"レポート生成完了: {self.output_excel}")

    def _create_summary_sheet(self, writer):
        """サマリーシート作成"""
        data = []

        for company in self.companies:
            analysis = self.analyses[company]
            latest_data = analysis['data'].iloc[-1]

            data.append({
                '銘柄': company,
                '最新年度': int(latest_data['年度']),
                '最新配当金': round(latest_data['配当金額'], 2),
                '配当利回り(%)': round(analysis['dividend_yield'][-1], 2),
                '配当性向(%)': round(analysis['payout_ratio'][-1], 2),
                '配当CAGR(%)': round(analysis['cagr'], 2),
                '連続増配年数': analysis['years_of_increase'],
                '持続可能性スコア': analysis['sustainability_score'],
                '株価': round(latest_data['株価'], 2),
                'EPS': round(latest_data['EPS'], 2)
            })

        summary_df = pd.DataFrame(data)
        summary_df.to_excel(writer, sheet_name='サマリー', index=False)

    def _create_history_sheet(self, writer):
        """配当履歴分析シート作成"""
        all_data = []

        for company in self.companies:
            analysis = self.analyses[company]
            company_data = analysis['data'].copy()
            company_data['配当利回り'] = analysis['dividend_yield']
            company_data['配当性向'] = analysis['payout_ratio']
            company_data['配当成長率(%)'] = analysis['dividend_growth']

            all_data.append(company_data[['銘柄', '年度', '配当金額', 'EPS', '株価',
                                          '配当利回り', '配当性向', '配当成長率(%)']])

        history_df = pd.concat(all_data, ignore_index=True)
        history_df = history_df.sort_values(['銘柄', '年度'])
        history_df.to_excel(writer, sheet_name='配当履歴分析', index=False)

    def _create_dgm_sheet(self, writer):
        """DGM評価シート作成"""
        data = []

        for company in self.companies:
            analysis = self.analyses[company]
            ddm = analysis['ddm_valuation']

            for scenario, values in ddm.items():
                row = {
                    '銘柄': company,
                    'シナリオ': scenario,
                    '理論株価': values['理論株価'],
                    '現在株価': values['現在株価'],
                    'アップサイド(%)': values['アップサイド']
                }
                data.append(row)

        dgm_df = pd.DataFrame(data)
        dgm_df.to_excel(writer, sheet_name='DGM評価', index=False)

    def _create_drip_sheet(self, writer):
        """再投資シミュレーションシート作成"""
        data = []

        for company in self.companies:
            analysis = self.analyses[company]
            drip = analysis['drip_simulation']

            for scenario, values in drip.items():
                row = {
                    '銘柄': company,
                    'シナリオ': scenario,
                    '株数': values['株数'],
                    '評価額': values['評価額'],
                    '初期投資比': values['初期投資比']
                }
                data.append(row)

        drip_df = pd.DataFrame(data)
        drip_df.to_excel(writer, sheet_name='再投資シミュレーション', index=False)

    def _create_sustainability_sheet(self, writer):
        """持続可能性シート作成"""
        data = []

        for company in self.companies:
            analysis = self.analyses[company]
            score = analysis['sustainability_score']

            # スコアに基づく評価
            if score >= 80:
                rating = '非常に安全'
            elif score >= 60:
                rating = '比較的安全'
            elif score >= 40:
                rating = '中程度のリスク'
            elif score >= 20:
                rating = '高リスク'
            else:
                rating = '非常に高リスク'

            data.append({
                '銘柄': company,
                '持続可能性スコア': score,
                '評価': rating,
                '平均配当性向(%)': round(np.mean(analysis['payout_ratio']), 2),
                '最新配当利回り(%)': round(analysis['dividend_yield'][-1], 2),
                '連続増配年数': analysis['years_of_increase']
            })

        sustainability_df = pd.DataFrame(data)
        sustainability_df.to_excel(writer, sheet_name='持続可能性', index=False)

    def _apply_styles(self):
        """Excelに書式を適用"""
        from openpyxl import load_workbook

        wb = load_workbook(self.output_excel)

        # ヘッダーのスタイル
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # ヘッダー行にスタイルを適用
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 列幅を自動調整
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

        wb.save(self.output_excel)

    def _add_charts(self):
        """グラフを追加"""
        # グラフ生成用の一時ファイルパスを作成
        charts_dir = self.output_excel.parent / 'charts'
        charts_dir.mkdir(exist_ok=True)

        for idx, company in enumerate(self.companies):
            analysis = self.analyses[company]
            company_data = analysis['data'].sort_values('年度')

            # グラフ1: 配当トレンドライン
            fig, axes = plt.subplots(2, 2, figsize=(14, 10))
            fig.suptitle(f'{company} - 配当分析チャート', fontsize=16, fontweight='bold')

            # 1. 配当金トレンド
            axes[0, 0].plot(company_data['年度'], company_data['配当金額'],
                           marker='o', linewidth=2, markersize=8, color='#2E75B6')
            axes[0, 0].set_title('配当金トレンド', fontweight='bold')
            axes[0, 0].set_xlabel('年度')
            axes[0, 0].set_ylabel('配当金額')
            axes[0, 0].grid(True, alpha=0.3)

            # 2. 利回り対株価
            scatter = axes[0, 1].scatter(company_data['株価'], analysis['dividend_yield'],
                                        s=100, alpha=0.6, c=range(len(company_data)),
                                        cmap='viridis')
            axes[0, 1].set_title('利回り対株価', fontweight='bold')
            axes[0, 1].set_xlabel('株価')
            axes[0, 1].set_ylabel('配当利回り(%)')
            axes[0, 1].grid(True, alpha=0.3)
            plt.colorbar(scatter, ax=axes[0, 1], label='年度')

            # 3. 配当性向推移
            axes[1, 0].bar(company_data['年度'], analysis['payout_ratio'],
                          color='#70AD47', alpha=0.7)
            axes[1, 0].axhline(y=50, color='red', linestyle='--', label='50% 警告線')
            axes[1, 0].axhline(y=70, color='darkred', linestyle='--', label='70% 危険線')
            axes[1, 0].set_title('配当性向推移', fontweight='bold')
            axes[1, 0].set_xlabel('年度')
            axes[1, 0].set_ylabel('配当性向(%)')
            axes[1, 0].legend()
            axes[1, 0].grid(True, alpha=0.3, axis='y')

            # 4. DRIP成長予測
            scenarios = ['10年後', '20年後', '30年後']
            multiples = [analysis['drip_simulation'][s]['初期投資比'] for s in scenarios]
            colors = ['#4472C4', '#70AD47', '#FFC000']
            axes[1, 1].bar(scenarios, multiples, color=colors, alpha=0.7)
            axes[1, 1].set_title('DRIP成長予測（初期投資比）', fontweight='bold')
            axes[1, 1].set_ylabel('初期投資に対する倍率')
            axes[1, 1].grid(True, alpha=0.3, axis='y')

            plt.tight_layout()
            chart_path = charts_dir / f'{idx+1:02d}_{company}_analysis.png'
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()

        print(f"チャート生成完了: {charts_dir} に {len(self.companies)} 個のチャートを保存")


def main():
    """メイン関数"""
    parser = argparse.ArgumentParser(
        description='配当分析スクリプト - 配当データを分析してExcelレポートを生成します'
    )
    parser.add_argument('input_csv', help='入力CSVファイルのパス')
    parser.add_argument('output_excel', help='出力Excelファイルのパス')

    args = parser.parse_args()

    # 分析実行
    analyzer = DividendAnalyzer(args.input_csv)
    analyzer.analyze()
    analyzer.generate_excel_report(args.output_excel)

    print("\n完了！すべての分析が完了しました。")
    print(f"出力ファイル: {analyzer.output_excel}")


if __name__ == '__main__':
    main()
