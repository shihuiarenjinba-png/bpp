#!/usr/bin/env python3
"""M&A pro forma EPS (accretion / dilution) → Excel report."""

from __future__ import annotations

import argparse
import itertools
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REQUIRED_FIELDS = (
    "ni_acquirer",
    "shares_acquirer",
    "ni_target",
    "shares_target",
    "ownership_pct",
    "consideration_cash",
    "consideration_shares",
    "new_debt_principal",
    "debt_interest_rate",
    "tax_rate",
)

OPTIONAL_NUMERIC = (
    "synergy_pretax",
    "synergy_aftertax",
    "one_time_costs_aftertax",
    "preferred_dividend",
)


def load_inputs(path: Path) -> dict[str, Any]:
    df = pd.read_csv(path)
    cols = {c.lower().strip(): c for c in df.columns}
    fcol = cols.get("field") or cols.get("key") or list(df.columns)[0]
    vcol = cols.get("value") or list(df.columns)[1]
    raw = dict(zip(df[fcol].astype(str).str.strip(), df[vcol]))
    out: dict[str, Any] = {}
    for k, v in raw.items():
        key = str(k).strip()
        if pd.isna(v) or v == "":
            out[key] = 0.0 if key in REQUIRED_FIELDS or key in OPTIONAL_NUMERIC else v
            continue
        try:
            out[key] = float(v)
        except (TypeError, ValueError):
            out[key] = v
    return out


def to_float(d: dict[str, Any], key: str, default: float = 0.0) -> float:
    v = d.get(key, default)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    return float(v)


def synergy_after_tax(d: dict[str, Any]) -> float:
    tax = to_float(d, "tax_rate", 0.30)
    after = to_float(d, "synergy_aftertax", 0.0)
    if after != 0.0:
        return after
    pre = to_float(d, "synergy_pretax", 0.0)
    return pre * (1.0 - tax) if pre else 0.0


def compute_core(d: dict[str, Any]) -> dict[str, Any]:
    ni_a = to_float(d, "ni_acquirer")
    sh_a = to_float(d, "shares_acquirer")
    ni_t = to_float(d, "ni_target")
    own = to_float(d, "ownership_pct", 1.0)
    tax = to_float(d, "tax_rate", 0.30)

    syn = synergy_after_tax(d)
    debt = to_float(d, "new_debt_principal")
    rate = to_float(d, "debt_interest_rate")
    interest_after_tax = debt * rate * (1.0 - tax)
    one_time = to_float(d, "one_time_costs_aftertax")
    pref = to_float(d, "preferred_dividend")

    new_shares = to_float(d, "consideration_shares")
    shares_pf = sh_a + new_shares

    if sh_a <= 0 or shares_pf <= 0:
        raise ValueError("shares_acquirer and pro forma shares must be positive.")

    ni_pf = ni_a + ni_t * own + syn - interest_after_tax - one_time - pref
    eps_a = ni_a / sh_a
    eps_pf = ni_pf / shares_pf
    if abs(eps_a) < 1e-12:
        acc_pct = float("nan")
    else:
        acc_pct = (eps_pf - eps_a) / abs(eps_a)

    return {
        "ni_acquirer": ni_a,
        "shares_acquirer": sh_a,
        "eps_standalone": eps_a,
        "ni_target_contrib": ni_t * own,
        "synergy_aftertax_used": syn,
        "interest_after_tax": interest_after_tax,
        "one_time_costs_aftertax": one_time,
        "preferred_dividend": pref,
        "ni_pro_forma": ni_pf,
        "shares_pro_forma": shares_pf,
        "eps_pro_forma": eps_pf,
        "accretion_dilution_pct": acc_pct,
    }


def sensitivity_grid(base: dict[str, Any], core: dict[str, Any]) -> pd.DataFrame:
    """2D grid: synergy multiplier × rate bump (annual)."""
    tax = to_float(base, "tax_rate", 0.30)
    debt = to_float(base, "new_debt_principal")
    base_rate = to_float(base, "debt_interest_rate")
    base_syn = synergy_after_tax(base)
    ni_a = to_float(base, "ni_acquirer")
    sh_a = to_float(base, "shares_acquirer")
    ni_t = to_float(base, "ni_target")
    own = to_float(base, "ownership_pct", 1.0)
    one_time = to_float(base, "one_time_costs_aftertax")
    pref = to_float(base, "preferred_dividend")
    new_shares = to_float(base, "consideration_shares")
    shares_pf = sh_a + new_shares

    syn_mults = (-0.2, 0.0, 0.2)
    rate_bumps = (-0.01, 0.0, 0.01)
    rows = []
    for sm, rb in itertools.product(syn_mults, rate_bumps):
        syn = base_syn * (1.0 + sm) if base_syn != 0 else 0.0
        rate = max(0.0, base_rate + rb)
        int_at = debt * rate * (1.0 - tax)
        ni_pf = ni_a + ni_t * own + syn - int_at - one_time - pref
        eps_pf = ni_pf / shares_pf
        eps_a = ni_a / sh_a
        acc = (eps_pf - eps_a) / abs(eps_a) if abs(eps_a) > 1e-12 else float("nan")
        rows.append(
            {
                "synergy_scenario": f"{int(sm * 100):+d}% vs base",
                "rate_bump_bp": int(rb * 10000),
                "eps_pro_forma": eps_pf,
                "accretion_dilution_pct": acc,
            }
        )
    return pd.DataFrame(rows)


def style_header(cell: Any) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell: Any) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_summary(ws: Any, core: dict[str, Any]) -> None:
    ws.title = "Summary"
    data = [
        ("Standalone EPS", core["eps_standalone"], "per share"),
        ("Pro forma EPS", core["eps_pro_forma"], "per share"),
        (
            "Accretion / (Dilution)",
            core["accretion_dilution_pct"],
            "fraction (e.g. 0.05 = +5%)",
        ),
        ("Pro forma shares", core["shares_pro_forma"], "shares"),
        ("Pro forma net income", core["ni_pro_forma"], "currency"),
    ]
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["C1"] = "Unit"
    for c in range(1, 4):
        style_header(ws.cell(1, c))
    for i, (name, val, unit) in enumerate(data, start=2):
        ws.cell(i, 1, name)
        if "Accretion" in name and val == val:  # not NaN
            ws.cell(i, 2, round(val * 100, 4))
            ws.cell(i, 3, "% (relative to standalone)")
        else:
            ws.cell(i, 2, round(val, 8) if isinstance(val, float) and val == val else val)
            ws.cell(i, 3, unit)
        for c in range(1, 4):
            style_cell(ws.cell(i, c))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28


def write_bridge(ws: Any, core: dict[str, Any]) -> None:
    ws.title = "Bridge"
    ws["A1"] = "Line item"
    ws["B1"] = "Amount"
    style_header(ws["A1"])
    style_header(ws["B1"])
    lines = [
        ("Acquirer net income (standalone)", core["ni_acquirer"]),
        ("(+) Target NI × ownership", core["ni_target_contrib"]),
        ("(+) After-tax synergies", core["synergy_aftertax_used"]),
        ("(−) After-tax interest on new debt", -core["interest_after_tax"]),
        ("(−) One-time after-tax costs", -core["one_time_costs_aftertax"]),
        ("(−) Preferred dividends", -core["preferred_dividend"]),
        ("= Pro forma net income", core["ni_pro_forma"]),
    ]
    for i, (label, amt) in enumerate(lines, start=2):
        ws.cell(i, 1, label)
        ws.cell(i, 2, round(amt, 6))
        style_cell(ws.cell(i, 1))
        style_cell(ws.cell(i, 2))
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22


def write_assumptions(ws: Any, d: dict[str, Any], limitations: list[str]) -> None:
    ws.title = "Assumptions"
    ws["A1"] = "field"
    ws["B1"] = "value"
    ws["C1"] = "notes"
    for c in range(1, 4):
        style_header(ws.cell(1, c))
    keys = sorted(set(d.keys()) | set(REQUIRED_FIELDS) | set(OPTIONAL_NUMERIC))
    row = 2
    for k in keys:
        ws.cell(row, 1, k)
        v = d.get(k, "")
        ws.cell(row, 2, v if v != "" else "")
        ws.cell(row, 3, "")
        for c in range(1, 4):
            style_cell(ws.cell(row, c))
        row += 1
    row += 1
    ws.cell(row, 1, "Limitations (not modeled)")
    style_header(ws.cell(row, 1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    for lim in limitations:
        ws.cell(row, 1, f"• {lim}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        style_cell(ws.cell(row, 1))
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40


def write_sensitivity(ws: Any, grid: pd.DataFrame) -> None:
    ws.title = "Sensitivity"
    for j, col in enumerate(grid.columns, start=1):
        c = ws.cell(1, j, col)
        style_header(c)
    for i, row in enumerate(grid.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j)
            if grid.columns[j - 1] == "accretion_dilution_pct" and val == val:
                cell.value = round(val * 100, 4)
            elif isinstance(val, float) and val == val:
                cell.value = round(val, 8)
            else:
                cell.value = val
            style_cell(cell)
    ws.cell(1, len(grid.columns) + 2, "Note:")
    ws.cell(
        2,
        len(grid.columns) + 2,
        "accretion_dilution_pct shown as % (×100). synergy_scenario is vs loaded synergy.",
    )
    for c in range(1, len(grid.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


def build_workbook(d: dict[str, Any], core: dict[str, Any], grid: pd.DataFrame) -> Workbook:
    wb = Workbook()
    write_summary(wb.active, core)
    write_bridge(wb.create_sheet(), core)
    limitations = [
        "Goodwill amortization / PPA step-up not modeled.",
        "Complex dilutives (options, converts) not modeled unless user adjusts shares manually.",
        "FX and consolidation mechanics simplified or omitted.",
    ]
    write_assumptions(wb.create_sheet(), d, limitations)
    write_sensitivity(wb.create_sheet(), grid)
    return wb


def validate_inputs(d: dict[str, Any]) -> None:
    missing = [f for f in REQUIRED_FIELDS if f not in d or d[f] == ""]
    if missing:
        raise ValueError(f"Missing required fields: {', '.join(missing)}")
    if to_float(d, "shares_acquirer") <= 0:
        raise ValueError("shares_acquirer must be positive.")


def main() -> int:
    p = argparse.ArgumentParser(description="M&A accretion / dilution Excel report")
    p.add_argument("input_csv", type=Path, help="ma_inputs.csv (field,value columns)")
    p.add_argument("output_xlsx", type=Path, help="Output .xlsx path")
    args = p.parse_args()

    try:
        d = load_inputs(args.input_csv)
        validate_inputs(d)
        core = compute_core(d)
        grid = sensitivity_grid(d, core)
        wb = build_workbook(d, core, grid)
        args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(args.output_xlsx)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Wrote {args.output_xlsx.resolve()}")
    print(
        f"Standalone EPS: {core['eps_standalone']:.6f} → Pro forma: {core['eps_pro_forma']:.6f} "
        f"({'accretion' if core['eps_pro_forma'] > core['eps_standalone'] else 'dilution'})"
    )
    if core["accretion_dilution_pct"] == core["accretion_dilution_pct"]:
        print(f"Accretion/(dilution): {core['accretion_dilution_pct'] * 100:.4f}%")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
