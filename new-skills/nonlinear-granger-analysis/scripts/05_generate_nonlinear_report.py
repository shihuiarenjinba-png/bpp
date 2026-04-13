"""
05_generate_nonlinear_report.py
================================
非線形グレンジャー分析の統合Excelレポート生成
4つの分析（サブサンプル・TAR・カーネル・ローリング）の結果を統合。
"""

import argparse, os, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

CH = "1F4E79"; FG = "FFFFFF"
C1 = "C0392B"; C2 = "E67E22"; C3 = "27AE60"; C4 = "2980B9"
ALT = "EBF5FB"; LG = "C6EFCE"; NR = "FFCCCC"

def hdr(ws, row, ncols, color=CH):
    for c in range(1, ncols+1):
        cl = ws.cell(row, c)
        cl.font  = Font(bold=True, color=FG, name="Calibri", size=10)
        cl.fill  = PatternFill(fgColor=color, fill_type="solid")
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def border(ws, r1, r2, c1, c2):
    b = Side(style="thin", color="B0B0B0")
    for row in ws.iter_rows(r1, r2, c1, c2):
        for cell in row:
            cell.border = Border(top=b, left=b, right=b, bottom=b)

def title(ws, row, col, text, size=12):
    c = ws.cell(row, col, text)
    c.font = Font(bold=True, size=size, name="Calibri", color=CH)
    return c


def build_overview_sheet(wb, linear_df, sub_df, tar_df, ker_df, rol_df):
    """統合比較シート（論文の核心）"""
    ws = wb.create_sheet("①統合比較マトリックス")
    title(ws, 1, 1, "非線形グレンジャー分析 統合比較（線形 vs 非線形）", 13)
    ws["A2"] = "線形で非有意だったペアに非線形手法で再検定。★★= 非線形有意、★= 弱い証拠"
    ws["A2"].font = Font(size=9, italic=True, color="666666", name="Calibri")

    cols = ["マクロ変数","ファクター","線形GC_p値","サブサンプル最小p値","TAR体制p値","カーネルp値","ローリング有意率(%)","総合判定"]
    sr = 4
    for ci, col in enumerate(cols, 1): ws.cell(sr, ci, col)
    hdr(ws, sr, len(cols))

    factors = ["Mkt-RF", "SMB", "HML", "RMW", "CMA"]
    macro   = ["CPI_growth","IP_growth","TERM_SPREAD","DEF_SPREAD","FX_return","OIL_return","VIX"]

    for ri, (mv, f) in enumerate([(mv,f) for mv in macro for f in factors], sr+1):
        ws.cell(ri, 1, mv).font = Font(name="Calibri", size=10)
        ws.cell(ri, 2, f).font  = Font(name="Calibri", size=10)

        # 線形p値
        lin_p = np.nan
        if linear_df is not None:
            row_l = linear_df[(linear_df["マクロ変数"]==mv) & (linear_df["ファクター"]==f)]
            if len(row_l)>0: lin_p = row_l.iloc[0].get("p値", np.nan)
        cell_l = ws.cell(ri, 3, round(float(lin_p),4) if not pd.isna(lin_p) else "N/A")
        if not pd.isna(lin_p):
            if float(lin_p) < 0.0014:
                cell_l.fill = PatternFill(fgColor=C1, fill_type="solid")
                cell_l.font = Font(color=FG, bold=True, name="Calibri")
            elif float(lin_p) < 0.05:
                cell_l.fill = PatternFill(fgColor=C2, fill_type="solid")
        cell_l.alignment = Alignment(horizontal="center")

        # サブサンプル最小p値
        sub_p = np.nan
        if sub_df is not None:
            row_s = sub_df[(sub_df["マクロ変数"]==mv) & (sub_df["ファクター"]==f)]
            if len(row_s)>0:
                p_cols = [c for c in sub_df.columns if "_p値" in c and c != "全サンプル_p値"]
                vals = [row_s.iloc[0].get(c) for c in p_cols
                        if not pd.isna(row_s.iloc[0].get(c, np.nan)) and isinstance(row_s.iloc[0].get(c), float)]
                if vals: sub_p = min(vals)
        ws.cell(ri, 4, round(float(sub_p),4) if not pd.isna(sub_p) else "N/A").alignment = Alignment(horizontal="center")
        if not pd.isna(sub_p) and float(sub_p) < 0.05:
            ws.cell(ri, 4).fill = PatternFill(fgColor=LG, fill_type="solid")

        # TAR体制p値
        tar_p = np.nan
        if tar_df is not None:
            row_t = tar_df[(tar_df["マクロ変数"]==mv) & (tar_df["ファクター"]==f)]
            if len(row_t)>0: tar_p = row_t.iloc[0].get("最小体制p値", np.nan)
        ws.cell(ri, 5, round(float(tar_p),4) if not pd.isna(tar_p) else "N/A").alignment = Alignment(horizontal="center")
        if not pd.isna(tar_p) and float(tar_p) < 0.05:
            ws.cell(ri, 5).fill = PatternFill(fgColor=LG, fill_type="solid")

        # カーネルp値
        ker_p = np.nan
        if ker_df is not None:
            row_k = ker_df[(ker_df["マクロ変数"]==mv) & (ker_df["ファクター"]==f)]
            if len(row_k)>0: ker_p = row_k.iloc[0].get("置換p値", np.nan)
        ws.cell(ri, 6, round(float(ker_p),4) if not pd.isna(ker_p) else "N/A").alignment = Alignment(horizontal="center")
        if not pd.isna(ker_p) and float(ker_p) < 0.05:
            ws.cell(ri, 6).fill = PatternFill(fgColor=LG, fill_type="solid")

        # ローリング有意率
        rol_rate = np.nan
        if rol_df is not None:
            row_r = rol_df[(rol_df["マクロ変数"]==mv) & (rol_df["ファクター"]==f)]
            if len(row_r)>0: rol_rate = row_r.iloc[0].get("有意率(%)", np.nan)
        ws.cell(ri, 7, round(float(rol_rate),1) if not pd.isna(rol_rate) else "N/A").alignment = Alignment(horizontal="center")
        if not pd.isna(rol_rate) and float(rol_rate) > 20:
            ws.cell(ri, 7).fill = PatternFill(fgColor=LG, fill_type="solid")

        # 総合判定
        scores = sum([
            (not pd.isna(sub_p) and float(sub_p) < 0.05) if not pd.isna(sub_p) else False,
            (not pd.isna(tar_p) and float(tar_p) < 0.05) if not pd.isna(tar_p) else False,
            (not pd.isna(ker_p) and float(ker_p) < 0.05) if not pd.isna(ker_p) else False,
            (not pd.isna(rol_rate) and float(rol_rate) > 20) if not pd.isna(rol_rate) else False,
        ])
        lin_nonsig = pd.isna(lin_p) or float(lin_p) >= 0.05
        judgment = ""
        if scores >= 3 and lin_nonsig: judgment = "★★ 強い非線形因果性"
        elif scores >= 2 and lin_nonsig: judgment = "★ 弱い非線形因果性"
        elif scores >= 1 and lin_nonsig: judgment = "△ 要追加検証"
        elif not lin_nonsig: judgment = "◎ 線形でも有意"
        else: judgment = "○ 真に非有意"
        ws.cell(ri, 8, judgment)
        if "★★" in judgment: ws.cell(ri, 8).fill = PatternFill(fgColor=C1, fill_type="solid"); ws.cell(ri, 8).font = Font(bold=True, color=FG, name="Calibri")
        elif "★ " in judgment: ws.cell(ri, 8).fill = PatternFill(fgColor=C2, fill_type="solid"); ws.cell(ri, 8).font = Font(bold=True, name="Calibri")

        if ri%2==1:
            for c in range(1, len(cols)+1):
                cc = ws.cell(ri, c)
                if cc.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cc.fill = PatternFill(fgColor=ALT, fill_type="solid")

    border(ws, sr, sr+35, 1, len(cols))
    for i, w in enumerate([18,10,12,16,12,12,16,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sheets_from_df(wb, df, sheet_name, title_text, note):
    if df is None or len(df)==0: return
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=12, name="Calibri", color=CH)
    ws["A2"] = note
    ws["A2"].font = Font(size=9, italic=True, color="666666", name="Calibri")
    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    hdr(ws, 4, len(df.columns))
    border(ws, 4, 4+len(df), 1, len(df.columns))
    for i in range(1, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def build_interpretation_sheet(wb):
    ws = wb.create_sheet("⑤解釈ガイド")
    ws["A1"] = "非線形グレンジャー分析 解釈ガイド（論文執筆用）"
    ws["A1"].font = Font(bold=True, size=13, name="Calibri", color=CH)

    content = [
        ("", ""),
        ("▌ 4手法の使い分け", ""),
        ("サブサンプル分析", "最も解釈しやすい。危機期に有意 → 「危機駆動型因果性」として論文に記載"),
        ("閾値グレンジャー(TAR)", "閾値推定値の経済的解釈が可能。DEF_SPREAD閾値 → 信用危機の臨界点"),
        ("カーネルグレンジャー", "最も一般的な非線形。計算負荷高。付録での報告が現実的"),
        ("ローリングウィンドウ", "時変因果性の可視化。図として論文本文に挿入すると説得力が高い"),
        ("", ""),
        ("▌ 結果パターン別の論文記述", ""),
        ("★★ 強い非線形因果性", "「線形グレンジャー検定では非有意（p=xxx）であったが、体制転換分析では危機期に有意（p=xxx）であり、非線形・条件付き因果関係の存在が示唆される」"),
        ("★ 弱い非線形因果性", "「頑健性チェックとして実施した非線形検定では、[手法]においてp<0.05を確認したが、他の手法では支持されず、結果の解釈には慎重を期す必要がある」"),
        ("○ 真に非有意", "「4種の非線形検定においても有意な因果性は検出されず、[マクロ変数X]と[ファクターY]の間には月次データにおける予測可能な先行関係は存在しないと結論づける」"),
        ("", ""),
        ("▌ 注意事項", ""),
        ("サンプルサイズ", "危機期サブサンプルはn=10〜20と小さく、統計的検出力が低い。結果の解釈は慎重に。"),
        ("多重検定", "4手法×35ペア=140通りの検定。FDR補正を適用すること。"),
        ("計算コスト", "カーネル検定はn_permutations=199で約10分。論文投稿前は999を推奨。"),
    ]

    cr = 3
    for heading, body in content:
        if heading.startswith("▌"):
            ws.cell(cr, 1, heading).font = Font(bold=True, size=11, name="Calibri", color=CH)
            cr += 1
        elif heading and body:
            ws.cell(cr, 1, heading).font = Font(bold=True, name="Calibri")
            ws.cell(cr, 2, body)
            ws.cell(cr, 2).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[cr].height = max(30, 15 + body.count("\n")*14)
            cr += 1
        else:
            cr += 1 if not heading and not body else 0
            cr += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--subsample", default="results/subsample_results.csv")
    parser.add_argument("--tar",       default="results/tar_results.csv")
    parser.add_argument("--kernel",    default="results/kernel_results.csv")
    parser.add_argument("--rolling",   default="results/rolling_results.csv")
    parser.add_argument("--linear",    default="results/granger_results.csv")
    parser.add_argument("--output",    default="outputs/nonlinear_analysis_report.xlsx")
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)

    def load(path):
        if os.path.exists(path):
            df = pd.read_csv(path, encoding="utf-8-sig", index_col=False)
            print(f"  読込: {path} ({len(df)}行)")
            return df
        print(f"  ⚠️ ファイル未発見: {path}")
        return None

    linear_df = load(args.linear)
    sub_df    = load(args.subsample)
    tar_df    = load(args.tar)
    ker_df    = load(args.kernel)
    rol_df    = load(args.rolling)

    wb = Workbook(); wb.remove(wb.active)
    build_overview_sheet(wb, linear_df, sub_df, tar_df, ker_df, rol_df)
    build_sheets_from_df(wb, sub_df, "②サブサンプル分析", "サブサンプル・グレンジャー検定（体制別p値）",
                          "危機期/平常期/低金利期/高金利期 各体制のp値と観測数")
    build_sheets_from_df(wb, tar_df, "③閾値グレンジャー(TAR)", "閾値グレンジャー因果性検定",
                          "閾値変数による体制分割。「閾値効果★」= 全体非有意→体制別有意")
    build_sheets_from_df(wb, ker_df, "④カーネルグレンジャー", "カーネルベース非線形グレンジャー検定（置換検定）",
                          "Diks-Panchenko型。置換p値 < 0.05 → 非線形因果性の存在証拠")
    build_sheets_from_df(wb, rol_df, "④ローリングウィンドウ", "ローリングウィンドウ・グレンジャー検定（時変因果性）",
                          "有意率(%)が高いペア → 時変的・断続的な因果性の存在")
    build_interpretation_sheet(wb)

    wb.save(args.output)
    print(f"\n✓ 非線形分析レポート保存: {args.output}")
    print("シート構成:")
    print("  ① 統合比較マトリックス（線形 vs 非線形）")
    print("  ② サブサンプル分析")
    print("  ③ 閾値グレンジャー（TAR）")
    print("  ④ カーネルグレンジャー / ローリングウィンドウ")
    print("  ⑤ 解釈ガイド（論文執筆用）")


if __name__ == "__main__":
    main()
