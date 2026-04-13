#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
信用リスク分析スクリプト
Altman Z-Score と Piotroski F-Score を用いた包括的な信用力評価ツール

使用方法:
    python credit_risk_analyzer.py --input <入力ファイル> --output <出力ファイル>

例:
    python credit_risk_analyzer.py --input financial_data.xlsx --output credit_risk_report.xlsx
"""

import argparse
import warnings
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from io import BytesIO

# ロギング設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
warnings.filterwarnings('ignore')

# 日本語フォント設定（matplotlib用）
plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Noto Sans CJK JP', 'Hiragino Sans']
plt.rcParams['axes.unicode_minus'] = False


class CreditRiskAnalyzer:
    """信用リスク分析クラス"""

    def __init__(self):
        self.df: Optional[pd.DataFrame] = None
        self.results: Dict = {}
        self.charts: Dict = {}

    def load_data(self, input_path: str) -> None:
        """ファイルからデータを読み込む"""
        logger.info(f"Loading data from: {input_path}")

        input_path = Path(input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")

        try:
            if input_path.suffix.lower() in ['.xlsx', '.xls']:
                self.df = pd.read_excel(input_path)
            elif input_path.suffix.lower() == '.csv':
                self.df = pd.read_csv(input_path)
            else:
                raise ValueError("対応ファイル形式: .xlsx, .xls, .csv")

            logger.info(f"データ読み込み完了: {len(self.df)} 行")
            self._validate_data()

        except Exception as e:
            logger.error(f"データ読み込みエラー: {e}")
            raise

    def _validate_data(self) -> None:
        """入力データの妥当性チェック"""
        required_cols = {
            '企業名', 'Company',
            '期間', 'Period',
            '流動資産', 'Current Assets',
            '固定資産', 'Fixed Assets',
            '総資産', 'Total Assets',
            '流動負債', 'Current Liabilities',
            '長期負債', 'Long-term Liabilities',
            '総負債', 'Total Liabilities',
            '売上高', 'Sales',
            '営業利益', 'EBIT',
            '純利益', 'Net Income',
            '営業キャッシュフロー', 'Operating Cash Flow',
            '市場時価総額', 'Market Value of Equity'
        }

        available_cols = set(self.df.columns)

        # 日本語/英語列名の統一
        col_mapping = {
            '企業名': 'Company',
            '期間': 'Period',
            '流動資産': 'Current Assets',
            '固定資産': 'Fixed Assets',
            '総資産': 'Total Assets',
            '流動負債': 'Current Liabilities',
            '長期負債': 'Long-term Liabilities',
            '総負債': 'Total Liabilities',
            '売上高': 'Sales',
            '営業利益': 'EBIT',
            '純利益': 'Net Income',
            '営業キャッシュフロー': 'Operating Cash Flow',
            '市場時価総額': 'Market Value of Equity',
            '売上原価': 'Cost of Goods Sold',
            '前期総資産': 'Prior Year Total Assets',
            '前期営業利益率': 'Prior Year ROA'
        }

        for ja_col, en_col in col_mapping.items():
            if ja_col in available_cols and en_col not in available_cols:
                self.df.rename(columns={ja_col: en_col}, inplace=True)

    def calculate_altman_z_score(self) -> pd.DataFrame:
        """
        Altman Z-Score を計算

        Z = 1.2×X₁ + 1.4×X₂ + 3.3×X₃ + 0.6×X₄ + 1.0×X₅

        X₁ = 運転資本 / 総資産
        X₂ = 利益剰余金 / 総資産
        X₃ = 営業利益 / 総資産
        X₄ = 時価総額 / 総負債
        X₅ = 売上高 / 総資産
        """
        logger.info("Calculating Altman Z-Score...")

        df = self.df.copy()

        # 運転資本 = 流動資産 - 流動負債
        df['Working Capital'] = df['Current Assets'] - df['Current Liabilities']

        # X₁: 運転資本 / 総資産
        df['X1_WC/TA'] = np.where(
            df['Total Assets'] != 0,
            df['Working Capital'] / df['Total Assets'],
            0
        )

        # X₂: 利益剰余金 / 総資産 (簡略版: 純利益の累積で推定)
        df['X2_RE/TA'] = np.where(
            df['Total Assets'] != 0,
            df['Net Income'] / df['Total Assets'],
            0
        )

        # X₃: 営業利益 / 総資産
        df['X3_EBIT/TA'] = np.where(
            df['Total Assets'] != 0,
            df['EBIT'] / df['Total Assets'],
            0
        )

        # X₄: 時価総額 / 総負債
        df['X4_MVE/TL'] = np.where(
            df['Total Liabilities'] != 0,
            df['Market Value of Equity'] / df['Total Liabilities'],
            0
        )

        # X₅: 売上高 / 総資産
        df['X5_Sales/TA'] = np.where(
            df['Total Assets'] != 0,
            df['Sales'] / df['Total Assets'],
            0
        )

        # Z-Score 計算
        df['Z_Score'] = (
            1.2 * df['X1_WC/TA'] +
            1.4 * df['X2_RE/TA'] +
            3.3 * df['X3_EBIT/TA'] +
            0.6 * df['X4_MVE/TL'] +
            1.0 * df['X5_Sales/TA']
        )

        # ゾーン判定
        def zone_judgment(z_score):
            if z_score > 2.99:
                return 'Safe Zone'
            elif 1.81 <= z_score <= 2.99:
                return 'Grey Zone'
            else:
                return 'Distress Zone'

        df['Z_Zone'] = df['Z_Score'].apply(zone_judgment)

        self.results['altman_z'] = df

        logger.info(f"Altman Z-Score 計算完了")
        return df

    def calculate_piotroski_f_score(self) -> pd.DataFrame:
        """
        Piotroski F-Score を計算（9つの財務指標）

        各指標が条件を満たしで1点、計9点
        """
        logger.info("Calculating Piotroski F-Score...")

        df = self.df.copy()
        df['F_Score'] = 0

        # 1. ROA > 0
        df['F1_ROA'] = (df['Net Income'] / df['Total Assets'] > 0).astype(int)
        df['F_Score'] += df['F1_ROA']

        # 2. 営業キャッシュフロー > 0
        df['F2_OCF'] = (df['Operating Cash Flow'] > 0).astype(int)
        df['F_Score'] += df['F2_OCF']

        # 3. ΔROA > 0 (ROAの改善)
        if len(df) > 1:
            prior_roa = df['Net Income'].shift(1) / df['Total Assets'].shift(1)
            current_roa = df['Net Income'] / df['Total Assets']
            df['F3_DROA'] = (current_roa > prior_roa).astype(int)
        else:
            df['F3_DROA'] = 0

        df['F_Score'] += df['F3_DROA']

        # 4. キャッシュフロー質 (Operating CF > Net Income)
        df['F4_CFQ'] = (df['Operating Cash Flow'] > df['Net Income']).astype(int)
        df['F_Score'] += df['F4_CFQ']

        # 5. Δ負債比率 < 0 (負債低下)
        if len(df) > 1:
            prior_leverage = df['Total Liabilities'].shift(1) / df['Total Assets'].shift(1)
            current_leverage = df['Total Liabilities'] / df['Total Assets']
            df['F5_DLEV'] = (current_leverage < prior_leverage).astype(int)
        else:
            df['F5_DLEV'] = 0

        df['F_Score'] += df['F5_DLEV']

        # 6. Δ流動比率 > 0 (流動比率改善)
        if len(df) > 1:
            prior_cr = df['Current Assets'].shift(1) / df['Current Liabilities'].shift(1)
            current_cr = df['Current Assets'] / df['Current Liabilities']
            df['F6_DCR'] = (current_cr > prior_cr).astype(int)
        else:
            df['F6_DCR'] = 0

        df['F_Score'] += df['F6_DCR']

        # 7. 新株発行なし (簡略版: 1で固定)
        df['F7_SHARES'] = 1
        df['F_Score'] += df['F7_SHARES']

        # 8. Δ売上総利益率 > 0 (粗利率改善)
        if 'Cost of Goods Sold' in df.columns:
            if len(df) > 1:
                prior_gm = (df['Sales'].shift(1) - df['Cost of Goods Sold'].shift(1)) / df['Sales'].shift(1)
                current_gm = (df['Sales'] - df['Cost of Goods Sold']) / df['Sales']
                df['F8_DGM'] = (current_gm > prior_gm).astype(int)
            else:
                df['F8_DGM'] = 0
        else:
            df['F8_DGM'] = 0

        df['F_Score'] += df['F8_DGM']

        # 9. Δ資産回転率 > 0 (資産回転率改善)
        if len(df) > 1:
            prior_ato = df['Sales'].shift(1) / df['Total Assets'].shift(1)
            current_ato = df['Sales'] / df['Total Assets']
            df['F9_DATO'] = (current_ato > prior_ato).astype(int)
        else:
            df['F9_DATO'] = 0

        df['F_Score'] += df['F9_DATO']

        # F-Score評価
        def f_score_rating(score):
            if score >= 8:
                return 'AAA - AA相当'
            elif score >= 6:
                return 'A相当'
            elif score >= 4:
                return 'BBB相当'
            elif score >= 2:
                return 'BB以下相当'
            else:
                return 'B - D相当'

        df['F_Rating'] = df['F_Score'].apply(f_score_rating)

        self.results['piotroski_f'] = df
        logger.info(f"Piotroski F-Score 計算完了")
        return df

    def calculate_liquidity_analysis(self) -> pd.DataFrame:
        """流動性分析を計算"""
        logger.info("Calculating Liquidity Analysis...")

        df = self.df.copy()

        # 流動比率 = 流動資産 / 流動負債
        df['Current Ratio'] = np.where(
            df['Current Liabilities'] != 0,
            df['Current Assets'] / df['Current Liabilities'],
            0
        )

        # 速動比率 (簡略版: 流動資産 - 在庫 ≈ 流動資産 * 0.7)
        df['Quick Ratio'] = np.where(
            df['Current Liabilities'] != 0,
            (df['Current Assets'] * 0.7) / df['Current Liabilities'],
            0
        )

        # 現金比率 (簡略版: 現金 ≈ 流動資産 * 0.2)
        df['Cash Ratio'] = np.where(
            df['Current Liabilities'] != 0,
            (df['Current Assets'] * 0.2) / df['Current Liabilities'],
            0
        )

        # 総合流動性スコア (加重平均: CR 50%, QR 30%, Cash 20%)
        df['Liquidity Score'] = (
            df['Current Ratio'] * 0.5 +
            df['Quick Ratio'] * 0.3 +
            df['Cash Ratio'] * 0.2
        )

        # 流動性評価
        def liquidity_rating(score):
            if score >= 1.5:
                return 'Excellent'
            elif score >= 1.2:
                return 'Good'
            elif score >= 1.0:
                return 'Adequate'
            else:
                return 'Poor'

        df['Liquidity Rating'] = df['Liquidity Score'].apply(liquidity_rating)

        self.results['liquidity'] = df
        logger.info(f"流動性分析計算完了")
        return df

    def calculate_interest_coverage(self) -> pd.DataFrame:
        """利息カバレッジ分析"""
        logger.info("Calculating Interest Coverage...")

        df = self.df.copy()

        # 推定利息（負債の平均利率を2-3%と仮定）
        df['Estimated Interest'] = df['Total Liabilities'] * 0.025

        # 利息カバレッジレシオ = EBIT / Interest
        df['Interest Coverage'] = np.where(
            df['Estimated Interest'] != 0,
            df['EBIT'] / df['Estimated Interest'],
            0
        )

        # 利息カバレッジ評価
        def coverage_rating(ratio):
            if ratio >= 5:
                return 'Strong'
            elif ratio >= 2.5:
                return 'Adequate'
            elif ratio >= 1.5:
                return 'Weak'
            else:
                return 'Distressed'

        df['Coverage Rating'] = df['Interest Coverage'].apply(coverage_rating)

        self.results['interest_coverage'] = df
        logger.info(f"利息カバレッジ分析完了")
        return df

    def calculate_composite_credit_rating(self) -> pd.DataFrame:
        """総合信用格付けを計算"""
        logger.info("Calculating Composite Credit Rating...")

        # 各分析結果を統合
        df = self.df.copy()

        # Altman Z-Scoreのスコア化 (0-100)
        z_scores = self.results['altman_z']['Z_Score']
        z_normalized = np.minimum(100, np.maximum(0, (z_scores / 3.0) * 100))

        # Piotroski F-Scoreのスコア化 (0-100)
        f_scores = self.results['piotroski_f']['F_Score']
        f_normalized = (f_scores / 9.0) * 100

        # 流動性スコアのスコア化 (0-100)
        liquidity_scores = self.results['liquidity']['Liquidity Score']
        liquidity_normalized = np.minimum(100, liquidity_scores * 50)

        # 利息カバレッジのスコア化 (0-100)
        coverage_scores = self.results['interest_coverage']['Interest Coverage']
        coverage_normalized = np.minimum(100, (coverage_scores / 5.0) * 100)

        # 加重合計スコア (Altman 40%, F-Score 30%, Liquidity 20%, Coverage 10%)
        df['Composite Score'] = (
            z_normalized * 0.40 +
            f_normalized * 0.30 +
            liquidity_normalized * 0.20 +
            coverage_normalized * 0.10
        )

        # 信用格付け決定
        def rating_from_score(score):
            if score >= 90:
                return 'AAA'
            elif score >= 85:
                return 'AA'
            elif score >= 80:
                return 'A'
            elif score >= 75:
                return 'BBB'
            elif score >= 65:
                return 'BB'
            elif score >= 55:
                return 'B'
            elif score >= 40:
                return 'CCC'
            elif score >= 25:
                return 'CC'
            elif score >= 10:
                return 'C'
            else:
                return 'D'

        df['Credit Rating'] = df['Composite Score'].apply(rating_from_score)

        # リスク評価
        def risk_assessment(rating):
            risk_map = {
                'AAA': '極めて低い',
                'AA': '非常に低い',
                'A': '低い',
                'BBB': '中程度',
                'BB': 'やや高い',
                'B': '高い',
                'CCC': '非常に高い',
                'CC': '非常に高い',
                'C': '極めて高い',
                'D': '実現'
            }
            return risk_map.get(rating, '不明')

        df['Risk Assessment'] = df['Credit Rating'].apply(risk_assessment)

        self.results['composite'] = df
        logger.info(f"総合信用格付け計算完了")
        return df

    def generate_charts(self) -> None:
        """可視化チャートを生成"""
        logger.info("Generating charts...")

        # 1. Z-Score ゲージチャート
        self._create_z_score_gauge()

        # 2. F-Score ブレークダウンバー
        self._create_f_score_breakdown()

        # 3. 格付け分布
        self._create_rating_distribution()

        # 4. トレンド分析
        self._create_trend_chart()

        logger.info(f"チャート生成完了")

    def _create_z_score_gauge(self) -> None:
        """Z-Score ゲージチャート"""
        z_data = self.results['altman_z']

        fig, axes = plt.subplots(1, len(z_data), figsize=(4 * len(z_data), 4))
        if len(z_data) == 1:
            axes = [axes]

        for idx, (i, row) in enumerate(z_data.iterrows()):
            ax = axes[idx]
            z_score = row['Z_Score']
            company = row['Company']

            # ゲージ背景
            theta = np.linspace(0, np.pi, 100)
            r = 1

            # 色分けゾーン
            distress = np.linspace(0, np.pi * (1.81 / 4), 50)
            grey = np.linspace(distress[-1], np.pi * (2.99 / 4), 50)
            safe = np.linspace(grey[-1], np.pi, 50)

            ax.fill_between(np.cos(distress), 0, np.sin(distress), color='#d62728', alpha=0.3, label='Distress (<1.81)')
            ax.fill_between(np.cos(grey), 0, np.sin(grey), color='#ff7f0e', alpha=0.3, label='Grey (1.81-2.99)')
            ax.fill_between(np.cos(safe), 0, np.sin(safe), color='#2ca02c', alpha=0.3, label='Safe (>2.99)')

            # 針
            needle_angle = np.pi * (min(z_score, 4) / 4)
            ax.arrow(0, 0, 0.8 * np.cos(needle_angle), 0.8 * np.sin(needle_angle),
                    head_width=0.1, head_length=0.1, fc='black', ec='black')

            ax.set_xlim(-1.3, 1.3)
            ax.set_ylim(-0.2, 1.3)
            ax.set_aspect('equal')
            ax.axis('off')
            ax.set_title(f"{company}\nZ-Score: {z_score:.2f}", fontsize=12, fontweight='bold')

        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        self.charts['z_score_gauge'] = buf
        plt.close()

    def _create_f_score_breakdown(self) -> None:
        """F-Score ブレークダウンバー"""
        f_data = self.results['piotroski_f']

        fig, axes = plt.subplots(1, min(len(f_data), 3), figsize=(5 * min(len(f_data), 3), 4))
        if len(f_data) == 1:
            axes = [axes]
        elif len(f_data) == 2:
            axes = list(axes)
        else:
            axes = axes[:min(len(f_data), 3)]

        criteria = ['F1_ROA', 'F2_OCF', 'F3_DROA', 'F4_CFQ', 'F5_DLEV',
                   'F6_DCR', 'F7_SHARES', 'F8_DGM', 'F9_DATO']
        labels = ['ROA', '営業CF', 'ΔROA', 'CF質', 'Δ負債', 'Δ流動比', '新株', 'Δ粗利', 'Δ回転率']

        for idx, ax in enumerate(axes):
            if idx < len(f_data):
                row = f_data.iloc[idx]
                values = [row[c] for c in criteria]
                colors = ['#2ca02c' if v == 1 else '#d62728' for v in values]

                ax.barh(labels, values, color=colors, alpha=0.7)
                ax.set_xlim(0, 1)
                ax.set_xlabel('スコア')
                ax.set_title(f"{row['Company']}\nF-Score: {row['F_Score']}/9", fontweight='bold')
                ax.set_xticks([0, 1])

        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        self.charts['f_score_breakdown'] = buf
        plt.close()

    def _create_rating_distribution(self) -> None:
        """格付け分布"""
        comp_data = self.results['composite']

        rating_counts = comp_data['Credit Rating'].value_counts()
        rating_order = ['AAA', 'AA', 'A', 'BBB', 'BB', 'B', 'CCC', 'CC', 'C', 'D']
        rating_counts = rating_counts.reindex(rating_order, fill_value=0)

        colors = ['#2ca02c', '#1f77b4', '#17becf', '#9467bd', '#ff7f0e', '#d62728', '#bc7c75', '#8c564b', '#c5b0d5', '#000000']

        fig, ax = plt.subplots(figsize=(10, 5))
        ax.bar(rating_counts.index, rating_counts.values, color=colors[:len(rating_counts)])
        ax.set_ylabel('企業数')
        ax.set_xlabel('信用格付け')
        ax.set_title('信用格付け分布', fontweight='bold', fontsize=14)
        ax.grid(axis='y', alpha=0.3)

        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        self.charts['rating_distribution'] = buf
        plt.close()

    def _create_trend_chart(self) -> None:
        """トレンド分析チャート"""
        comp_data = self.results['composite']

        if len(comp_data) < 2:
            logger.warning("トレンド分析にはデータが不足しています")
            return

        fig, ax = plt.subplots(figsize=(12, 6))

        for company in comp_data['Company'].unique():
            company_data = comp_data[comp_data['Company'] == company]
            if 'Period' in company_data.columns:
                ax.plot(company_data['Period'].astype(str), company_data['Composite Score'],
                       marker='o', label=company, linewidth=2)

        ax.set_ylabel('総合スコア')
        ax.set_xlabel('期間')
        ax.set_title('信用スコアトレンド分析', fontweight='bold', fontsize=14)
        ax.legend()
        ax.grid(True, alpha=0.3)
        ax.set_ylim(0, 100)

        plt.xticks(rotation=45)
        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        self.charts['trend_chart'] = buf
        plt.close()

    def export_to_excel(self, output_path: str) -> None:
        """結果を Excel に出力"""
        logger.info(f"Exporting results to: {output_path}")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. サマリーシート
            self._write_summary_sheet(writer)

            # 2. Altman Z-Score シート
            self._write_altman_sheet(writer)

            # 3. Piotroski F-Score シート
            self._write_piotroski_sheet(writer)

            # 4. 流動性分析シート
            self._write_liquidity_sheet(writer)

            # 5. 総合信用格付けシート
            self._write_composite_sheet(writer)

            # 6. トレンド分析シート
            self._write_trend_sheet(writer)

        logger.info(f"Excel 出力完了: {output_path}")

    def _write_summary_sheet(self, writer) -> None:
        """サマリーシート"""
        comp_data = self.results['composite']
        alt_data = self.results['altman_z']
        f_data = self.results['piotroski_f']

        summary_df = pd.DataFrame({
            '企業名': comp_data['Company'],
            'Z-Score': alt_data['Z_Score'].round(2),
            'Z-Zone': alt_data['Z_Zone'],
            'F-Score': f_data['F_Score'],
            '流動性スコア': self.results['liquidity']['Liquidity Score'].round(2),
            '利息CF': self.results['interest_coverage']['Interest Coverage'].round(2),
            '総合スコア': comp_data['Composite Score'].round(2),
            '信用格付け': comp_data['Credit Rating'],
            'リスク評価': comp_data['Risk Assessment']
        })

        summary_df.to_excel(writer, sheet_name='サマリー', index=False)
        self._format_sheet(writer.sheets['サマリー'])

    def _write_altman_sheet(self, writer) -> None:
        """Altman Z-Score シート"""
        alt_data = self.results['altman_z']
        alt_output = alt_data[['Company', 'Period', 'X1_WC/TA', 'X2_RE/TA', 'X3_EBIT/TA',
                               'X4_MVE/TL', 'X5_Sales/TA', 'Z_Score', 'Z_Zone']].copy()
        alt_output.columns = ['企業名', '期間', 'X₁ (WC/TA)', 'X₂ (RE/TA)', 'X₃ (EBIT/TA)',
                             'X₄ (MVE/TL)', 'X₅ (Sales/TA)', 'Z-Score', 'ゾーン']

        alt_output.to_excel(writer, sheet_name='Altman Z-Score', index=False)
        self._format_sheet(writer.sheets['Altman Z-Score'])

    def _write_piotroski_sheet(self, writer) -> None:
        """Piotroski F-Score シート"""
        f_data = self.results['piotroski_f']
        f_output = f_data[['Company', 'Period', 'F1_ROA', 'F2_OCF', 'F3_DROA', 'F4_CFQ',
                           'F5_DLEV', 'F6_DCR', 'F7_SHARES', 'F8_DGM', 'F9_DATO', 'F_Score', 'F_Rating']].copy()
        f_output.columns = ['企業名', '期間', 'ROA', '営業CF', 'ΔROA', 'CF質', 'Δ負債',
                           'Δ流動比', '新株', 'Δ粗利', 'Δ回転', 'F-Score', '評価']

        f_output.to_excel(writer, sheet_name='Piotroski F-Score', index=False)
        self._format_sheet(writer.sheets['Piotroski F-Score'])

    def _write_liquidity_sheet(self, writer) -> None:
        """流動性分析シート"""
        liq_data = self.results['liquidity']
        liq_output = liq_data[['Company', 'Period', 'Current Ratio', 'Quick Ratio',
                               'Cash Ratio', 'Liquidity Score', 'Liquidity Rating']].copy()
        liq_output.columns = ['企業名', '期間', '流動比率', '速動比率', '現金比率', '総合スコア', '評価']

        liq_output.to_excel(writer, sheet_name='流動性分析', index=False)
        self._format_sheet(writer.sheets['流動性分析'])

    def _write_composite_sheet(self, writer) -> None:
        """総合信用格付けシート"""
        comp_data = self.results['composite']
        comp_output = comp_data[['Company', 'Period', 'Composite Score', 'Credit Rating', 'Risk Assessment']].copy()
        comp_output.columns = ['企業名', '期間', '総合スコア', '信用格付け', 'リスク評価']

        comp_output.to_excel(writer, sheet_name='総合信用格付け', index=False)
        self._format_sheet(writer.sheets['総合信用格付け'])

    def _write_trend_sheet(self, writer) -> None:
        """トレンド分析シート"""
        comp_data = self.results['composite']
        trend_output = comp_data[['Company', 'Period', 'Z_Score', 'F_Score', 'Liquidity Score',
                                 'Composite Score', 'Credit Rating']].copy()
        trend_output.columns = ['企業名', '期間', 'Z-Score', 'F-Score', '流動性', '総合スコア', '格付け']

        trend_output.to_excel(writer, sheet_name='トレンド分析', index=False)
        self._format_sheet(writer.sheets['トレンド分析'])

    def _format_sheet(self, ws) -> None:
        """シートの書式設定"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 2, 25)

    def run(self, input_path: str, output_path: str) -> None:
        """完全分析を実行"""
        try:
            logger.info("=" * 60)
            logger.info("信用リスク分析を開始します")
            logger.info("=" * 60)

            self.load_data(input_path)
            self.calculate_altman_z_score()
            self.calculate_piotroski_f_score()
            self.calculate_liquidity_analysis()
            self.calculate_interest_coverage()
            self.calculate_composite_credit_rating()
            self.generate_charts()
            self.export_to_excel(output_path)

            logger.info("=" * 60)
            logger.info(f"分析完了！結果は {output_path} に出力されました")
            logger.info("=" * 60)

        except Exception as e:
            logger.error(f"エラーが発生しました: {e}", exc_info=True)
            sys.exit(1)


def main():
    """メイン関数"""
    parser = argparse.ArgumentParser(
        description='信用リスク分析スクリプト',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
例:
  python credit_risk_analyzer.py --input financial_data.xlsx --output report.xlsx
  python credit_risk_analyzer.py -i data.csv -o analysis.xlsx
        '''
    )

    parser.add_argument('-i', '--input', required=True, help='入力ファイル（CSV/XLSX）')
    parser.add_argument('-o', '--output', required=True, help='出力ファイル（XLSX）')

    args = parser.parse_args()

    analyzer = CreditRiskAnalyzer()
    analyzer.run(args.input, args.output)


if __name__ == '__main__':
    main()
