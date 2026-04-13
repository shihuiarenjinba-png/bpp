#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
キャッシュフロー予測スクリプト
Cash Flow Forecasting Tool

このスクリプトは過去のキャッシュフロー実績から、
複数シナリオの将来予測と資金繰り分析を実行します。
"""

import argparse
import sys
import os
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import rcParams

# 日本語フォント設定
rcParams['font.sans-serif'] = ['DejaVu Sans']


def load_data(input_path):
    """CSVファイルからデータを読み込む"""
    try:
        df = pd.read_csv(input_path)
        df['日付'] = pd.to_datetime(df['日付'])
        df = df.sort_values('日付').reset_index(drop=True)
        return df
    except Exception as e:
        print(f"エラー: ファイル読み込み失敗 - {e}", file=sys.stderr)
        sys.exit(1)


def analyze_historical_trends(df):
    """過去のキャッシュフロー傾向を分析"""
    df = df.copy()
    df['合計CF'] = df['営業CF'] + df['投資CF'] + df['財務CF']

    # 3ヶ月移動平均
    df['3M移動平均'] = df['合計CF'].rolling(window=3, center=True).mean()

    # トレンド計算
    x = np.arange(len(df))
    z = np.polyfit(x, df['合計CF'].values, 1)
    trend = z[0]  # 傾き

    return df, trend


def detect_seasonality(df):
    """季節性パターンを検出"""
    df = df.copy()
    df['月'] = df['日付'].dt.month
    df['合計CF'] = df['営業CF'] + df['投資CF'] + df['財務CF']

    # 月別の平均を計算
    monthly_avg = df.groupby('月')['合計CF'].mean()
    overall_avg = df['合計CF'].mean()

    # 季節性指数 = 月別平均 / 全体平均
    seasonal_indices = {}
    for month in range(1, 13):
        if month in monthly_avg.index:
            seasonal_indices[month] = monthly_avg[month] / overall_avg if overall_avg != 0 else 1.0
        else:
            seasonal_indices[month] = 1.0

    return seasonal_indices


def create_forecast(df, seasonal_indices, trend, months_ahead=12):
    """3シナリオの予測を作成"""
    last_date = df['日付'].iloc[-1]
    last_cf = (df['営業CF'].iloc[-1] + df['投資CF'].iloc[-1] +
               df['財務CF'].iloc[-1])

    forecast_dates = []
    forecast_basic = []
    forecast_optimistic = []
    forecast_pessimistic = []

    # 直近3ヶ月の平均CF
    recent_cf = (df['営業CF'].iloc[-3:].mean() +
                 df['投資CF'].iloc[-3:].mean() +
                 df['財務CF'].iloc[-3:].mean())

    for i in range(1, months_ahead + 1):
        forecast_date = last_date + pd.DateOffset(months=i)
        forecast_dates.append(forecast_date)

        month = forecast_date.month
        seasonal_factor = seasonal_indices.get(month, 1.0)

        # トレンド調整
        trend_adjustment = trend * i

        # 基本シナリオ: 直近平均 × 季節性 + トレンド
        base_forecast = recent_cf * seasonal_factor + trend_adjustment

        # 楽観シナリオ: 基本 × 1.15
        optimistic = base_forecast * 1.15

        # 悲観シナリオ: 基本 × 0.85
        pessimistic = base_forecast * 0.85

        forecast_basic.append(base_forecast)
        forecast_optimistic.append(optimistic)
        forecast_pessimistic.append(pessimistic)

    forecast_df = pd.DataFrame({
        '日付': forecast_dates,
        '基本シナリオ': forecast_basic,
        '楽観シナリオ': forecast_optimistic,
        '悲観シナリオ': forecast_pessimistic
    })

    return forecast_df


def calculate_cumulative_position(df, forecast_df, initial_cash=10000000):
    """累積現金ポジションを計算"""
    df_analysis = df.copy()
    df_analysis['合計CF'] = (df_analysis['営業CF'] +
                            df_analysis['投資CF'] +
                            df_analysis['財務CF'])
    df_analysis['累積CF'] = df_analysis['合計CF'].cumsum()
    df_analysis['現金ポジション'] = initial_cash + df_analysis['累積CF']

    # 予測期間の現金ポジション
    forecast_cf = forecast_df.copy()
    last_position = df_analysis['現金ポジション'].iloc[-1]

    forecast_cf['基本累積CF'] = forecast_cf['基本シナリオ'].cumsum()
    forecast_cf['楽観累積CF'] = forecast_cf['楽観シナリオ'].cumsum()
    forecast_cf['悲観累積CF'] = forecast_cf['悲観シナリオ'].cumsum()

    forecast_cf['基本現金ポジション'] = last_position + forecast_cf['基本累積CF']
    forecast_cf['楽観現金ポジション'] = last_position + forecast_cf['楽観累積CF']
    forecast_cf['悲観現金ポジション'] = last_position + forecast_cf['悲観累積CF']

    return df_analysis, forecast_cf


def calculate_burn_rate(df):
    """バーンレート（月平均キャッシュ消費）を計算"""
    df = df.copy()
    df['合計CF'] = df['営業CF'] + df['投資CF'] + df['財務CF']

    # 直近3ヶ月の平均
    recent_cf = df['合計CF'].iloc[-3:].mean()
    burn_rate = max(0, -recent_cf) if recent_cf < 0 else 0

    return burn_rate


def detect_warning_flags(df_analysis, forecast_cf, initial_cash=10000000):
    """リスク警告フラグを検出"""
    warnings = []

    # 連続した負CF月の検出
    df = df_analysis.copy()
    df['合計CF'] = df['営業CF'] + df['投資CF'] + df['財務CF']
    negative_months = (df['合計CF'] < 0).astype(int)
    max_consecutive = 0
    current_consecutive = 0

    for val in negative_months:
        if val == 1:
            current_consecutive += 1
            max_consecutive = max(max_consecutive, current_consecutive)
        else:
            current_consecutive = 0

    if max_consecutive >= 3:
        warnings.append({
            'フラグ': '連続負CF',
            '重度': '高',
            '説明': f'{max_consecutive}ヶ月連続の負キャッシュフロー',
            'アクション': 'コスト構造見直し、売上拡大施策の実行'
        })

    # 現金ランウェイ計算（悲観シナリオ）
    burn_rate = calculate_burn_rate(df)

    if burn_rate > 0:
        last_position = df_analysis['現金ポジション'].iloc[-1]
        runway_months = last_position / burn_rate if burn_rate > 0 else float('inf')

        if runway_months < 6:
            warnings.append({
                'フラグ': 'ランウェイ不足',
                '重度': '高' if runway_months < 3 else '中',
                '説明': f'現金ランウェイが{runway_months:.1f}ヶ月',
                'アクション': '融資申請、資金調達の検討'
            })

    # 悲観シナリオで現金ショート可能性
    min_pessimistic = forecast_cf['悲観現金ポジション'].min()
    if min_pessimistic < 0:
        months_to_short = forecast_cf[forecast_cf['悲観現金ポジション'] < 0].index.min()
        warnings.append({
            'フラグ': '現金ショートリスク',
            '重度': '極',
            '説明': f'悲観シナリオで{months_to_short + 1}ヶ月目に現金がショート',
            'アクション': '即座に融資枠確保、経営施策の急速な改善'
        })

    # 季節性の大きな変動
    df['月'] = df['日付'].dt.month
    monthly_std = df.groupby('月')['合計CF'].std().mean()
    if monthly_std > df['合計CF'].mean() * 0.5:
        warnings.append({
            'フラグ': '高い季節変動',
            '重度': '中',
            '説明': '月次の変動が大きく、資金計画が複雑',
            'アクション': '季節対応の融資枠設定、在庫最適化'
        })

    return warnings


def generate_excel_report(output_path, df_analysis, forecast_df, forecast_cf,
                          warnings, initial_cash=10000000):
    """Excel形式のレポートを生成"""
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. サマリーシート
    ws_summary = wb.create_sheet("サマリー", 0)
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    row = 1
    ws_summary[f'A{row}'] = 'キャッシュフロー予測サマリー'
    ws_summary[f'A{row}'].font = title_font

    row = 3
    ws_summary[f'A{row}'] = '基本情報'
    ws_summary[f'A{row}'].font = subheader_font

    row = 4
    ws_summary[f'A{row}'] = '分析開始日'
    ws_summary[f'B{row}'] = df_analysis['日付'].iloc[0].strftime('%Y-%m-%d')
    row += 1
    ws_summary[f'A{row}'] = '分析終了日'
    ws_summary[f'B{row}'] = df_analysis['日付'].iloc[-1].strftime('%Y-%m-%d')
    row += 1
    ws_summary[f'A{row}'] = '分析期間'
    ws_summary[f'B{row}'] = f'{len(df_analysis)}ヶ月'
    row += 1

    row += 1
    ws_summary[f'A{row}'] = '主要指標'
    ws_summary[f'A{row}'].font = subheader_font

    row += 1
    last_position = df_analysis['現金ポジション'].iloc[-1]
    ws_summary[f'A{row}'] = '現在の現金ポジション'
    ws_summary[f'B{row}'] = f'{last_position:,.0f}円'
    row += 1

    burn_rate = calculate_burn_rate(df_analysis)
    ws_summary[f'A{row}'] = 'バーンレート（月平均）'
    ws_summary[f'B{row}'] = f'{burn_rate:,.0f}円/月'
    row += 1

    runway = (last_position / burn_rate) if burn_rate > 0 else float('inf')
    ws_summary[f'A{row}'] = 'キャッシュランウェイ'
    ws_summary[f'B{row}'] = f'{runway:.1f}ヶ月' if runway != float('inf') else '無限'
    row += 1

    avg_cf = df_analysis['合計CF'].mean()
    ws_summary[f'A{row}'] = '過去平均CF（月）'
    ws_summary[f'B{row}'] = f'{avg_cf:,.0f}円'
    row += 1

    row += 1
    ws_summary[f'A{row}'] = '12ヶ月予測結果'
    ws_summary[f'A{row}'].font = subheader_font

    row += 1
    ws_summary[f'A{row}'] = '基本シナリオ累積CF'
    ws_summary[f'B{row}'] = f'{forecast_cf["基本累積CF"].iloc[-1]:,.0f}円'
    row += 1
    ws_summary[f'A{row}'] = '楽観シナリオ累積CF'
    ws_summary[f'B{row}'] = f'{forecast_cf["楽観累積CF"].iloc[-1]:,.0f}円'
    row += 1
    ws_summary[f'A{row}'] = '悲観シナリオ累積CF'
    ws_summary[f'B{row}'] = f'{forecast_cf["悲観累積CF"].iloc[-1]:,.0f}円'
    row += 1

    row += 1
    ws_summary[f'A{row}'] = '12ヶ月後の予測現金ポジション'
    ws_summary[f'A{row}'].font = subheader_font

    row += 1
    ws_summary[f'A{row}'] = '基本シナリオ'
    ws_summary[f'B{row}'] = f'{forecast_cf["基本現金ポジション"].iloc[-1]:,.0f}円'
    row += 1
    ws_summary[f'A{row}'] = '楽観シナリオ'
    ws_summary[f'B{row}'] = f'{forecast_cf["楽観現金ポジション"].iloc[-1]:,.0f}円'
    row += 1
    ws_summary[f'A{row}'] = '悲観シナリオ'
    ws_summary[f'B{row}'] = f'{forecast_cf["悲観現金ポジション"].iloc[-1]:,.0f}円'
    row += 1

    # 2. 実績分析シート
    ws_history = wb.create_sheet("実績分析", 1)
    ws_history.column_dimensions['A'].width = 12
    ws_history.column_dimensions['B'].width = 15
    ws_history.column_dimensions['C'].width = 15
    ws_history.column_dimensions['D'].width = 15
    ws_history.column_dimensions['E'].width = 15

    headers = ['日付', '営業CF', '投資CF', '財務CF', '合計CF']
    for col, header in enumerate(headers, 1):
        cell = ws_history.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for r_idx, row_data in df_analysis.iterrows():
        ws_history.cell(row=r_idx+2, column=1).value = row_data['日付'].strftime('%Y-%m')
        ws_history.cell(row=r_idx+2, column=2).value = row_data['営業CF']
        ws_history.cell(row=r_idx+2, column=3).value = row_data['投資CF']
        ws_history.cell(row=r_idx+2, column=4).value = row_data['財務CF']
        ws_history.cell(row=r_idx+2, column=5).value = row_data['合計CF']

    # 3. 季節性パターンシート
    ws_seasonal = wb.create_sheet("季節性パターン", 2)
    ws_seasonal.column_dimensions['A'].width = 12
    ws_seasonal.column_dimensions['B'].width = 20

    seasonal_indices = detect_seasonality(df_analysis)
    ws_seasonal['A1'] = '月'
    ws_seasonal['B1'] = '季節性指数'
    ws_seasonal['A1'].fill = header_fill
    ws_seasonal['B1'].fill = header_fill
    ws_seasonal['A1'].font = header_font
    ws_seasonal['B1'].font = header_font

    for month in range(1, 13):
        ws_seasonal[f'A{month+1}'] = f'{month}月'
        ws_seasonal[f'B{month+1}'] = seasonal_indices.get(month, 1.0)

    # 4. シナリオ予測シート
    ws_forecast = wb.create_sheet("シナリオ予測", 3)
    ws_forecast.column_dimensions['A'].width = 12
    ws_forecast.column_dimensions['B'].width = 18
    ws_forecast.column_dimensions['C'].width = 18
    ws_forecast.column_dimensions['D'].width = 18

    headers_forecast = ['日付', '基本シナリオ', '楽観シナリオ', '悲観シナリオ']
    for col, header in enumerate(headers_forecast, 1):
        cell = ws_forecast.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for r_idx, row_data in forecast_df.iterrows():
        ws_forecast.cell(row=r_idx+2, column=1).value = row_data['日付'].strftime('%Y-%m')
        ws_forecast.cell(row=r_idx+2, column=2).value = row_data['基本シナリオ']
        ws_forecast.cell(row=r_idx+2, column=3).value = row_data['楽観シナリオ']
        ws_forecast.cell(row=r_idx+2, column=4).value = row_data['悲観シナリオ']

    # 5. 資金繰り表シート
    ws_cashflow = wb.create_sheet("資金繰り表", 4)
    ws_cashflow.column_dimensions['A'].width = 12
    ws_cashflow.column_dimensions['B'].width = 15
    ws_cashflow.column_dimensions['C'].width = 18
    ws_cashflow.column_dimensions['D'].width = 18
    ws_cashflow.column_dimensions['E'].width = 18

    headers_cf = ['日付', '合計CF', '基本現金', '楽観現金', '悲観現金']
    for col, header in enumerate(headers_cf, 1):
        cell = ws_cashflow.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 過去実績
    row_num = 2
    for r_idx, row_data in df_analysis.iterrows():
        ws_cashflow.cell(row=row_num, column=1).value = row_data['日付'].strftime('%Y-%m')
        ws_cashflow.cell(row=row_num, column=2).value = row_data['合計CF']
        ws_cashflow.cell(row=row_num, column=3).value = row_data['現金ポジション']
        row_num += 1

    # 予測
    for r_idx, row_data in forecast_cf.iterrows():
        ws_cashflow.cell(row=row_num, column=1).value = row_data['日付'].strftime('%Y-%m')
        ws_cashflow.cell(row=row_num, column=2).value = row_data['基本シナリオ']
        ws_cashflow.cell(row=row_num, column=3).value = row_data['基本現金ポジション']
        ws_cashflow.cell(row=row_num, column=4).value = row_data['楽観現金ポジション']
        ws_cashflow.cell(row=row_num, column=5).value = row_data['悲観現金ポジション']
        row_num += 1

    # 6. 警告フラグシート
    ws_warnings = wb.create_sheet("警告フラグ", 5)
    ws_warnings.column_dimensions['A'].width = 15
    ws_warnings.column_dimensions['B'].width = 10
    ws_warnings.column_dimensions['C'].width = 35
    ws_warnings.column_dimensions['D'].width = 40

    headers_warn = ['フラグ', '重度', '説明', 'アクション']
    for col, header in enumerate(headers_warn, 1):
        cell = ws_warnings.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    if warnings:
        for r_idx, warning in enumerate(warnings, 2):
            ws_warnings.cell(row=r_idx, column=1).value = warning['フラグ']
            ws_warnings.cell(row=r_idx, column=2).value = warning['重度']
            ws_warnings.cell(row=r_idx, column=3).value = warning['説明']
            ws_warnings.cell(row=r_idx, column=4).value = warning['アクション']
    else:
        ws_warnings.cell(row=2, column=1).value = 'リスク警告なし'

    wb.save(output_path)
    print(f"Excelレポート出力完了: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='キャッシュフロー予測スクリプト',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument('--input', required=True, help='入力CSVファイルパス')
    parser.add_argument('--output', required=True, help='出力Excelファイルパス')
    parser.add_argument('--initial-cash', type=float, default=10000000,
                       help='初期現金ポジション')
    parser.add_argument('--forecast-months', type=int, default=12,
                       help='予測期間（ヶ月）')

    args = parser.parse_args()

    # データ読み込み
    print("データを読み込み中...")
    df = load_data(args.input)

    if len(df) < 6:
        print("エラー: 最低6ヶ月のデータが必要です", file=sys.stderr)
        sys.exit(1)

    # 実績分析
    print("実績分析を実行中...")
    df_analysis, trend = analyze_historical_trends(df)

    # 季節性検出
    print("季節性を検出中...")
    seasonal_indices = detect_seasonality(df_analysis)

    # 予測作成
    print("シナリオ予測を生成中...")
    forecast_df = create_forecast(df_analysis, seasonal_indices, trend,
                                 args.forecast_months)

    # 現金ポジション計算
    print("現金ポジションを計算中...")
    df_analysis, forecast_cf = calculate_cumulative_position(
        df_analysis, forecast_df, args.initial_cash
    )

    # リスク検出
    print("リスク警告を検出中...")
    warnings = detect_warning_flags(df_analysis, forecast_cf, args.initial_cash)

    # Excelレポート生成
    print("Excelレポートを生成中...")
    generate_excel_report(args.output, df_analysis, forecast_df, forecast_cf,
                         warnings, args.initial_cash)

    # サマリー出力
    print("\n=== 分析完了 ===")
    print(f"現在の現金ポジション: {df_analysis['現金ポジション'].iloc[-1]:,.0f}円")
    print(f"バーンレート: {calculate_burn_rate(df_analysis):,.0f}円/月")
    print(f"キャッシュランウェイ: {(df_analysis['現金ポジション'].iloc[-1] / max(calculate_burn_rate(df_analysis), 1)):,.1f}ヶ月")
    print(f"検出された警告: {len(warnings)}件")

    if warnings:
        print("\n警告の詳細:")
        for w in warnings:
            print(f"  - {w['フラグ']} ({w['重度']}): {w['説明']}")


if __name__ == '__main__':
    main()
