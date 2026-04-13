#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比較財務分析スクリプト
複数企業の財務データを包括的に比較し、Excel形式のレポートを生成します。
"""

import argparse
import os
import glob
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib
from matplotlib import rcParams
import warnings

warnings.filterwarnings('ignore')

# 日本語フォント設定
matplotlib.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'DejaVu Sans', 'Noto Sans CJK JP']
matplotlib.rcParams['axes.unicode_minus'] = False


class ComparativeFinancialAnalyzer:
    """比較財務分析クラス"""

    def __init__(self, company_column='会社名'):
        self.company_column = company_column
        self.df = None
        self.companies = []
        self.summary_stats = {}
        self.common_size_pl = None
        self.common_size_bs = None
        self.growth_rates = None
        self.financial_ratios = None

    def load_data(self, input_path):
        """
        入力ファイルまたはファイルグループからデータを読み込む

        Args:
            input_path: ファイルパス、複数ファイルパターン、またはディレクトリ
        """
        input_path = str(input_path)

        # ワイルドカード展開
        if '*' in input_path:
            files = glob.glob(input_path)
        elif os.path.isdir(input_path):
            files = glob.glob(os.path.join(input_path, '*.xlsx')) + \
                   glob.glob(os.path.join(input_path, '*.xls')) + \
                   glob.glob(os.path.join(input_path, '*.csv'))
        else:
            files = [input_path]

        if not files:
            raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")

        dfs = []
        for file_path in sorted(files):
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8-sig')
            else:
                df = pd.read_excel(file_path)
            dfs.append(df)

        # すべてのDataFrameを統合
        self.df = pd.concat(dfs, ignore_index=False, sort=False)

        # 会社名列でインデックスを設定
        if self.company_column in self.df.columns:
            self.df.set_index(self.company_column, inplace=True)

        # 数値列を抽出
        self.df = self.df.apply(pd.to_numeric, errors='coerce')
        self.companies = self.df.index.unique().tolist()

        print(f"データ読み込み完了: {len(self.companies)} 企業")
        print(f"企業: {', '.join(self.companies)}")

    def calculate_common_size_pl(self):
        """コモンサイズ損益計算書を計算（売上に対する百分率）"""
        revenue_cols = ['売上', 'Revenue', '売上高', '営業収益']
        revenue = None

        for col in revenue_cols:
            if col in self.df.columns:
                revenue = self.df[col]
                break

        if revenue is None:
            print("警告: 売上列が見つかりません")
            revenue = self.df.iloc[:, 0]

        # PL関連列を特定
        pl_items = ['売上原価', 'Cost of Goods Sold', '営業費用', 'Operating Expenses',
                    '営業利益', 'Operating Income', '純利益', 'Net Income', 'EBITDA']

        # 存在する列のみを抽出
        available_pl = [col for col in pl_items if col in self.df.columns]

        if not available_pl:
            available_pl = [col for col in self.df.columns if col != revenue.name]

        # コモンサイズ計算
        self.common_size_pl = pd.DataFrame()
        for col in available_pl:
            if col in self.df.columns:
                self.common_size_pl[col] = (self.df[col] / revenue * 100).round(2)

        # 売上を100%として追加
        self.common_size_pl['売上'] = 100.0

        return self.common_size_pl

    def calculate_common_size_bs(self):
        """コモンサイズ貸借対照表を計算（総資産に対する百分率）"""
        assets_cols = ['総資産', 'Total Assets', '資産合計']
        total_assets = None

        for col in assets_cols:
            if col in self.df.columns:
                total_assets = self.df[col]
                break

        if total_assets is None:
            print("警告: 総資産列が見つかりません")
            return None

        # BS関連列を特定
        bs_items = ['流動資産', 'Current Assets', '固定資産', 'Fixed Assets',
                    '流動負債', 'Current Liabilities', '固定負債', 'Long-term Liabilities',
                    '株主資本', 'Shareholders Equity', '負債合計', 'Total Liabilities']

        available_bs = [col for col in bs_items if col in self.df.columns]

        if not available_bs:
            available_bs = [col for col in self.df.columns
                          if col not in ['売上', 'Revenue', '営業利益', '純利益']]

        # コモンサイズ計算
        self.common_size_bs = pd.DataFrame()
        for col in available_bs:
            if col in self.df.columns:
                self.common_size_bs[col] = (self.df[col] / total_assets * 100).round(2)

        # 総資産を100%として追加
        self.common_size_bs['総資産'] = 100.0

        return self.common_size_bs

    def calculate_growth_rates(self):
        """成長率を計算（YoY およ CAGR）"""
        self.growth_rates = pd.DataFrame()

        # 数値列ごとに成長率を計算
        for col in self.df.columns:
            if col in ['売上', 'Revenue', '売上高']:
                # 各企業ごとにYoYを計算
                df_sorted = self.df.sort_index()
                for company in self.companies:
                    if company in df_sorted.index:
                        company_data = df_sorted.loc[company, col]
                        if len(company_data) >= 2:
                            yoy = ((company_data.iloc[-1] - company_data.iloc[-2]) /
                                  company_data.iloc[-2] * 100)
                            self.growth_rates.loc[company, f'{col}_YoY%'] = round(yoy, 2)

                            # CAGRを計算（3年以上のデータがある場合）
                            if len(company_data) >= 3:
                                years = len(company_data) - 1
                                cagr = ((company_data.iloc[-1] / company_data.iloc[0])
                                       ** (1/years) - 1) * 100
                                self.growth_rates.loc[company, f'{col}_CAGR%'] = round(cagr, 2)

        return self.growth_rates

    def calculate_financial_ratios(self):
        """主要財務指標を計算"""
        ratios = {}

        for company in self.companies:
            company_ratios = {}
            company_data = self.df.loc[[company]].iloc[0]

            # ROE（自己資本利益率）
            if '純利益' in company_data.index and '株主資本' in company_data.index:
                if company_data['株主資本'] > 0:
                    company_ratios['ROE%'] = round(
                        (company_data['純利益'] / company_data['株主資本']) * 100, 2)

            # ROA（総資産利益率）
            if '純利益' in company_data.index and '総資産' in company_data.index:
                if company_data['総資産'] > 0:
                    company_ratios['ROA%'] = round(
                        (company_data['純利益'] / company_data['総資産']) * 100, 2)

            # 営業利益率
            if '営業利益' in company_data.index and '売上' in company_data.index:
                if company_data['売上'] > 0:
                    company_ratios['営業利益率%'] = round(
                        (company_data['営業利益'] / company_data['売上']) * 100, 2)

            # 純利益率
            if '純利益' in company_data.index and '売上' in company_data.index:
                if company_data['売上'] > 0:
                    company_ratios['純利益率%'] = round(
                        (company_data['純利益'] / company_data['売上']) * 100, 2)

            # 総資産回転率
            if '売上' in company_data.index and '総資産' in company_data.index:
                if company_data['総資産'] > 0:
                    company_ratios['総資産回転率'] = round(
                        company_data['売上'] / company_data['総資産'], 2)

            # 流動比率
            if '流動資産' in company_data.index and '流動負債' in company_data.index:
                if company_data['流動負債'] > 0:
                    company_ratios['流動比率'] = round(
                        company_data['流動資産'] / company_data['流動負債'], 2)

            # 負債比率
            if '負債合計' in company_data.index and '総資産' in company_data.index:
                if company_data['総資産'] > 0:
                    company_ratios['負債比率%'] = round(
                        (company_data['負債合計'] / company_data['総資産']) * 100, 2)

            # 利息カバー率（営業利益 / 利息費用）
            if '営業利益' in company_data.index and '利息費用' in company_data.index:
                if company_data['利息費用'] > 0 and company_data['営業利益'] > 0:
                    company_ratios['利息カバー率'] = round(
                        company_data['営業利益'] / company_data['利息費用'], 2)

            ratios[company] = company_ratios

        self.financial_ratios = pd.DataFrame(ratios).T
        return self.financial_ratios

    def create_ranking_table(self):
        """各指標での順位付けテーブルを作成"""
        ranking = pd.DataFrame()

        if self.financial_ratios is not None:
            for col in self.financial_ratios.columns:
                ranking[f'{col}_順位'] = self.financial_ratios[col].rank(
                    method='min', ascending=False).astype(int)

        return ranking

    def evaluate_strengths_weaknesses(self):
        """企業ごとの強み・弱み分析"""
        evaluation = {}

        if self.financial_ratios is not None:
            mean_ratios = self.financial_ratios.mean()

            for company in self.companies:
                company_ratios = self.financial_ratios.loc[company]

                strengths = []
                weaknesses = []

                for metric in company_ratios.index:
                    if pd.notna(company_ratios[metric]) and pd.notna(mean_ratios[metric]):
                        if company_ratios[metric] > mean_ratios[metric] * 1.1:
                            strengths.append(f"{metric}: {company_ratios[metric]:.2f}")
                        elif company_ratios[metric] < mean_ratios[metric] * 0.9:
                            weaknesses.append(f"{metric}: {company_ratios[metric]:.2f}")

                evaluation[company] = {
                    '強み': ', '.join(strengths) if strengths else 'なし',
                    '弱み': ', '.join(weaknesses) if weaknesses else 'なし',
                    '総合スコア': round(
                        (company_ratios / mean_ratios).mean() * 100, 1)
                }

        return pd.DataFrame(evaluation).T

    def create_charts(self, output_dir):
        """分析チャートを作成"""
        output_dir = Path(output_dir).parent / 'charts'
        output_dir.mkdir(exist_ok=True)

        try:
            # グループ棒グラフ：主要指標比較
            if not self.financial_ratios.empty:
                fig, ax = plt.subplots(figsize=(12, 6))

                # 百分率指標のみを抽出
                pct_cols = [col for col in self.financial_ratios.columns
                           if '%' in col]
                if pct_cols:
                    self.financial_ratios[pct_cols].plot(kind='bar', ax=ax)
                    ax.set_title('主要財務指標比較（百分率）', fontsize=14, fontweight='bold')
                    ax.set_xlabel('企業名', fontsize=12)
                    ax.set_ylabel('パーセンテージ (%)', fontsize=12)
                    ax.legend(loc='upper left', fontsize=10)
                    ax.grid(axis='y', alpha=0.3)
                    plt.tight_layout()
                    plt.savefig(output_dir / 'main_metrics_comparison.png', dpi=300, bbox_inches='tight')
                    plt.close()
        except Exception as e:
            print(f"チャート作成エラー: {e}")

        try:
            # レーダーチャート
            if not self.financial_ratios.empty:
                from math import pi

                # 正規化データを作成（0-100スケール）
                normalized = pd.DataFrame()
                for col in self.financial_ratios.columns:
                    if '%' not in col:
                        min_val = self.financial_ratios[col].min()
                        max_val = self.financial_ratios[col].max()
                        if max_val > min_val:
                            normalized[col] = ((self.financial_ratios[col] - min_val) /
                                             (max_val - min_val) * 100)
                        else:
                            normalized[col] = 50

                if not normalized.empty and len(normalized.columns) >= 3:
                    fig, ax = plt.subplots(figsize=(10, 10),
                                         subplot_kw=dict(projection='polar'))

                    categories = list(normalized.columns)
                    N = len(categories)
                    angles = [n / float(N) * 2 * pi for n in range(N)]
                    angles += angles[:1]

                    colors = plt.cm.Set3(np.linspace(0, 1, len(self.companies)))

                    for idx, company in enumerate(self.companies):
                        if company in normalized.index:
                            values = normalized.loc[company].values.tolist()
                            values += values[:1]
                            ax.plot(angles, values, 'o-', linewidth=2,
                                   label=company, color=colors[idx])
                            ax.fill(angles, values, alpha=0.15, color=colors[idx])

                    ax.set_xticks(angles[:-1])
                    ax.set_xticklabels(categories, size=10)
                    ax.set_ylim(0, 100)
                    ax.set_title('企業総合評価レーダーチャート',
                               fontsize=14, fontweight='bold', pad=20)
                    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
                    ax.grid(True)

                    plt.tight_layout()
                    plt.savefig(output_dir / 'radar_chart.png', dpi=300, bbox_inches='tight')
                    plt.close()
        except Exception as e:
            print(f"レーダーチャート作成エラー: {e}")

        try:
            # 成長軌跡グラフ
            if self.growth_rates is not None and not self.growth_rates.empty:
                fig, ax = plt.subplots(figsize=(10, 6))

                growth_cols = [col for col in self.growth_rates.columns
                             if 'YoY' in col]
                if growth_cols:
                    self.growth_rates[growth_cols].plot(kind='bar', ax=ax)
                    ax.set_title('成長率比較（YoY）', fontsize=14, fontweight='bold')
                    ax.set_xlabel('企業名', fontsize=12)
                    ax.set_ylabel('成長率 (%)', fontsize=12)
                    ax.legend(loc='upper left', fontsize=10)
                    ax.axhline(y=0, color='black', linestyle='-', linewidth=0.5)
                    ax.grid(axis='y', alpha=0.3)
                    plt.tight_layout()
                    plt.savefig(output_dir / 'growth_trajectory.png', dpi=300, bbox_inches='tight')
                    plt.close()
        except Exception as e:
            print(f"成長軌跡チャート作成エラー: {e}")

        print(f"チャートを保存しました: {output_dir}")

    def export_to_excel(self, output_path):
        """結果をExcel形式で出力"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # サマリーシート
            summary_data = {
                '企業': self.companies,
            }

            if not self.financial_ratios.empty:
                for col in self.financial_ratios.columns:
                    summary_data[col] = self.financial_ratios[col].values

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='サマリー', index=False)

            # コモンサイズPL
            if self.common_size_pl is not None:
                self.common_size_pl.to_excel(writer, sheet_name='コモンサイズPL')

            # コモンサイズBS
            if self.common_size_bs is not None:
                self.common_size_bs.to_excel(writer, sheet_name='コモンサイズBS')

            # 成長率比較
            if self.growth_rates is not None and not self.growth_rates.empty:
                self.growth_rates.to_excel(writer, sheet_name='成長率比較')

            # 指標ランキング
            ranking = self.create_ranking_table()
            if not ranking.empty:
                ranking.to_excel(writer, sheet_name='指標ランキング')

            # 総合評価
            evaluation = self.evaluate_strengths_weaknesses()
            evaluation.to_excel(writer, sheet_name='総合評価')

        # Excelのスタイル調整
        self._format_excel(output_path)

        print(f"Excel レポートを出力しました: {output_path}")

    def _format_excel(self, excel_path):
        """Excelファイルのスタイルを整える"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_path)

            header_fill = PatternFill(start_color='366092', end_color='366092',
                                     fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet in wb.sheetnames:
                ws = wb[sheet]

                # ヘッダーのスタイル
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = border

                # データセルのスタイル
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00'

                # 列幅自動調整
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(excel_path)
        except Exception as e:
            print(f"Excel スタイル調整エラー: {e}")


def main():
    parser = argparse.ArgumentParser(
        description='比較財務分析スクリプト'
    )
    parser.add_argument('--input', '-i', required=True,
                       help='入力ファイルパス（CSV/Excel）またはワイルドカードパターン')
    parser.add_argument('--output', '-o', required=True,
                       help='出力Excelファイルパス')
    parser.add_argument('--company-column', '-c', default='会社名',
                       help='企業名列の名前（デフォルト: 会社名）')

    args = parser.parse_args()

    try:
        # アナライザーの初期化と実行
        analyzer = ComparativeFinancialAnalyzer(company_column=args.company_column)

        print("データ読み込み中...")
        analyzer.load_data(args.input)

        print("コモンサイズPL計算中...")
        analyzer.calculate_common_size_pl()

        print("コモンサイズBS計算中...")
        analyzer.calculate_common_size_bs()

        print("成長率計算中...")
        analyzer.calculate_growth_rates()

        print("財務指標計算中...")
        analyzer.calculate_financial_ratios()

        print("チャート作成中...")
        analyzer.create_charts(args.output)

        print("Excel出力中...")
        analyzer.export_to_excel(args.output)

        print("\n完了しました。")

    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
