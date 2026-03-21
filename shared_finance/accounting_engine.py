from __future__ import annotations

import io
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill


COLUMN_CANDIDATES = {
    "date": ["date", "booking_date", "transaction_date", "伝票日付", "日付", "年月日"],
    "description": ["description", "details", "memo", "摘要", "内容", "description_1"],
    "amount": ["amount", "金額", "value", "net_amount", "支払額", "入金額"],
    "debit": ["debit", "借方", "debit_amount"],
    "credit": ["credit", "貸方", "credit_amount"],
    "currency": ["currency", "通貨"],
    "account": ["account", "account_name", "勘定科目", "科目"],
    "category": ["category", "区分", "費目"],
    "counterparty": ["counterparty", "取引先", "vendor", "customer"],
}


ACCOUNT_RULES = [
    {
        "keywords": ["売上", "sales", "revenue", "income"],
        "account": "Sales Revenue",
        "bucket": "Revenue",
        "type": "revenue",
        "cash_flow": "Operating",
        "ifrs": "IFRS 15 Revenue",
        "confidence": 0.95,
    },
    {
        "keywords": ["仕入", "inventory", "purchase", "原価", "cogs"],
        "account": "Cost of Goods Sold",
        "bucket": "Cost of Sales",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "Cost of sales",
        "confidence": 0.9,
    },
    {
        "keywords": ["給料", "給与", "payroll", "salary", "bonus"],
        "account": "Payroll Expense",
        "bucket": "Payroll",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "Employee benefits expense",
        "confidence": 0.92,
    },
    {
        "keywords": ["家賃", "rent", "lease", "office"],
        "account": "Rent Expense",
        "bucket": "Rent",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "Lease or rent expense",
        "confidence": 0.86,
    },
    {
        "keywords": ["広告", "marketing", "ad", "promotion"],
        "account": "Marketing Expense",
        "bucket": "Marketing",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "Selling expense",
        "confidence": 0.84,
    },
    {
        "keywords": ["tax", "法人税", "消費税", "税"],
        "account": "Tax Expense",
        "bucket": "Tax",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "Tax expense",
        "confidence": 0.8,
    },
    {
        "keywords": ["借入", "loan", "debt", "interest", "利息"],
        "account": "Financing / Interest",
        "bucket": "Financing",
        "type": "financing",
        "cash_flow": "Financing",
        "ifrs": "Borrowings / finance cost",
        "confidence": 0.82,
    },
    {
        "keywords": ["設備", "capex", "machine", "asset", "software"],
        "account": "Capital Expenditure",
        "bucket": "Capex",
        "type": "capex",
        "cash_flow": "Investing",
        "ifrs": "Property, plant and equipment / intangible",
        "confidence": 0.8,
    },
    {
        "keywords": ["売掛", "receivable", "ar"],
        "account": "Accounts Receivable",
        "bucket": "Working Capital",
        "type": "working_capital",
        "cash_flow": "Non-cash / Working capital",
        "ifrs": "Trade receivables",
        "confidence": 0.78,
    },
    {
        "keywords": ["買掛", "payable", "ap"],
        "account": "Accounts Payable",
        "bucket": "Working Capital",
        "type": "working_capital",
        "cash_flow": "Non-cash / Working capital",
        "ifrs": "Trade payables",
        "confidence": 0.78,
    },
]


DEFAULT_BUCKET_ORDER = [
    "Revenue",
    "Cost of Sales",
    "Payroll",
    "Rent",
    "Marketing",
    "Tax",
    "Financing",
    "Capex",
    "Working Capital",
    "Other",
]


def _normalized_column_lookup(df: pd.DataFrame) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for col in df.columns:
        normalized = str(col).strip().lower().replace(" ", "_")
        lookup[normalized] = str(col)
    return lookup


def _pick_column(df: pd.DataFrame, logical_name: str) -> str | None:
    lookup = _normalized_column_lookup(df)
    for candidate in COLUMN_CANDIDATES[logical_name]:
        candidate_key = candidate.strip().lower().replace(" ", "_")
        if candidate_key in lookup:
            return lookup[candidate_key]

    for candidate in COLUMN_CANDIDATES[logical_name]:
        candidate_key = candidate.strip().lower()
        for normalized, original in lookup.items():
            if candidate_key in normalized:
                return original
    return None


def _parse_numeric(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("¥", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def _infer_rule(text: str) -> dict[str, object]:
    lowered = (text or "").lower()
    for rule in ACCOUNT_RULES:
        if any(keyword in lowered for keyword in rule["keywords"]):
            return rule
    return {
        "account": "Unmapped Transaction",
        "bucket": "Other",
        "type": "expense",
        "cash_flow": "Operating",
        "ifrs": "To be classified",
        "confidence": 0.35,
    }


def parse_tabular_upload(file_name: str, file_bytes: bytes) -> list[tuple[str, pd.DataFrame]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(io.BytesIO(file_bytes))
        return [("CSV", df)]
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        workbook = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
        return list(workbook.items())
    return []


def parse_manual_table(text: str) -> pd.DataFrame:
    if not text or not text.strip():
        return pd.DataFrame()
    sample = text.strip()
    sep = "\t" if "\t" in sample else ","
    try:
        return pd.read_csv(io.StringIO(sample), sep=sep)
    except Exception:
        return pd.DataFrame()


def maybe_ocr_image(file_bytes: bytes) -> str:
    try:
        from PIL import Image
        import pytesseract

        image = Image.open(io.BytesIO(file_bytes))
        return pytesseract.image_to_string(image, lang="jpn+eng").strip()
    except Exception:
        return ""


def normalize_transactions(
    frame: pd.DataFrame,
    *,
    source_name: str,
    currency: str,
) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame()

    df = frame.copy()
    date_col = _pick_column(df, "date")
    description_col = _pick_column(df, "description")
    amount_col = _pick_column(df, "amount")
    debit_col = _pick_column(df, "debit")
    credit_col = _pick_column(df, "credit")
    currency_col = _pick_column(df, "currency")
    account_col = _pick_column(df, "account")
    category_col = _pick_column(df, "category")
    counterparty_col = _pick_column(df, "counterparty")

    normalized = pd.DataFrame()
    normalized["booking_date"] = (
        pd.to_datetime(df[date_col], errors="coerce")
        if date_col
        else pd.NaT
    )
    normalized["booking_date"] = normalized["booking_date"].fillna(pd.Timestamp.today().normalize())

    if description_col:
        normalized["description"] = df[description_col].fillna("").astype(str)
    else:
        normalized["description"] = "Imported transaction"

    if amount_col:
        amount = _parse_numeric(df[amount_col])
    else:
        debit = _parse_numeric(df[debit_col]) if debit_col else pd.Series(0.0, index=df.index)
        credit = _parse_numeric(df[credit_col]) if credit_col else pd.Series(0.0, index=df.index)
        amount = debit.fillna(0.0) - credit.fillna(0.0)

    normalized["amount"] = amount.fillna(0.0)
    normalized["currency"] = (
        df[currency_col].fillna(currency).astype(str) if currency_col else currency
    )
    normalized["raw_account"] = (
        df[account_col].fillna("").astype(str) if account_col else ""
    )
    normalized["raw_category"] = (
        df[category_col].fillna("").astype(str) if category_col else ""
    )
    normalized["counterparty"] = (
        df[counterparty_col].fillna("").astype(str) if counterparty_col else ""
    )
    normalized["source_name"] = source_name

    descriptor = (
        normalized["description"].fillna("")
        + " "
        + normalized["raw_account"].fillna("")
        + " "
        + normalized["raw_category"].fillna("")
    )
    rules = descriptor.apply(_infer_rule)

    normalized["inferred_account"] = rules.apply(lambda item: item["account"])
    normalized["budget_bucket"] = rules.apply(lambda item: item["bucket"])
    normalized["bucket_type"] = rules.apply(lambda item: item["type"])
    normalized["cash_flow_class"] = rules.apply(lambda item: item["cash_flow"])
    normalized["ifrs_hint"] = rules.apply(lambda item: item["ifrs"])
    normalized["mapping_confidence"] = rules.apply(lambda item: item["confidence"])
    normalized["journal_direction"] = np.where(
        normalized["amount"] >= 0, "Debit / Inflow", "Credit / Outflow"
    )
    normalized["cash_movement"] = np.where(
        normalized["cash_flow_class"].eq("Non-cash / Working capital"),
        0.0,
        normalized["amount"],
    )

    normalized = normalized[
        normalized["description"].str.strip().ne("") | normalized["amount"].ne(0)
    ].copy()
    normalized["abs_amount"] = normalized["amount"].abs()
    normalized.sort_values(by="booking_date", inplace=True)
    normalized.reset_index(drop=True, inplace=True)
    normalized.insert(0, "tx_id", [f"TX-{idx:05d}" for idx in range(1, len(normalized) + 1)])
    return normalized


def demo_transactions() -> pd.DataFrame:
    demo = pd.DataFrame(
        {
            "Date": pd.date_range("2025-01-05", periods=8, freq="15D"),
            "Description": [
                "January sales settlement",
                "Office rent",
                "Payroll transfer",
                "Marketing campaign",
                "Inventory purchase",
                "Software asset",
                "Customer receivable",
                "Interest payment",
            ],
            "Amount": [2500000, -260000, -540000, -180000, -720000, -320000, 900000, -45000],
            "Category": [
                "Sales",
                "Rent",
                "Payroll",
                "Marketing",
                "Purchase",
                "Capex",
                "Receivable",
                "Interest",
            ],
        }
    )
    return demo


def build_budget_template(journal_df: pd.DataFrame) -> pd.DataFrame:
    if journal_df is None or journal_df.empty:
        rows = []
        for bucket in DEFAULT_BUCKET_ORDER:
            bucket_type = "revenue" if bucket == "Revenue" else "expense"
            rows.append(
                {
                    "budget_bucket": bucket,
                    "budget_type": bucket_type,
                    "monthly_budget": 0.0,
                    "annual_budget": 0.0,
                    "owner": "TBD",
                }
            )
        return pd.DataFrame(rows)

    working = journal_df.copy()
    grouped = working.groupby(["budget_bucket", "bucket_type"], as_index=False)["amount"].mean()
    grouped.rename(columns={"bucket_type": "budget_type"}, inplace=True)
    grouped["monthly_budget"] = np.where(
        grouped["budget_type"].eq("revenue"),
        grouped["amount"].clip(lower=0.0),
        grouped["amount"].abs(),
    )
    grouped["monthly_budget"] = grouped["monthly_budget"].round(0)
    grouped["annual_budget"] = grouped["monthly_budget"] * 12.0
    grouped["owner"] = "Finance"
    grouped = grouped[["budget_bucket", "budget_type", "monthly_budget", "annual_budget", "owner"]]

    known = grouped["budget_bucket"].tolist()
    for bucket in DEFAULT_BUCKET_ORDER:
        if bucket not in known:
            grouped.loc[len(grouped)] = {
                "budget_bucket": bucket,
                "budget_type": "revenue" if bucket == "Revenue" else "expense",
                "monthly_budget": 0.0,
                "annual_budget": 0.0,
                "owner": "Finance",
            }

    grouped.sort_values(
        by="budget_bucket",
        key=lambda series: series.map({bucket: idx for idx, bucket in enumerate(DEFAULT_BUCKET_ORDER)}).fillna(999),
        inplace=True,
    )
    grouped.reset_index(drop=True, inplace=True)
    return grouped


def build_budget_vs_actual(journal_df: pd.DataFrame, budget_df: pd.DataFrame) -> pd.DataFrame:
    if budget_df is None or budget_df.empty:
        return pd.DataFrame()

    if journal_df is None or journal_df.empty:
        result = budget_df.copy()
        result["actual_monthly"] = 0.0
        result["variance"] = result["monthly_budget"]
        result["variance_pct"] = 0.0
        return result

    actuals = journal_df.groupby(["budget_bucket", "bucket_type"], as_index=False)["amount"].mean()
    actuals.rename(columns={"bucket_type": "budget_type"}, inplace=True)
    actuals["actual_monthly"] = np.where(
        actuals["budget_type"].eq("revenue"),
        actuals["amount"].clip(lower=0.0),
        actuals["amount"].abs(),
    )
    merged = budget_df.merge(
        actuals[["budget_bucket", "actual_monthly"]],
        how="left",
        on="budget_bucket",
    )
    merged["actual_monthly"] = merged["actual_monthly"].fillna(0.0)
    merged["variance"] = merged["monthly_budget"] - merged["actual_monthly"]
    merged["variance_pct"] = np.where(
        merged["monthly_budget"].abs() > 1e-9,
        merged["variance"] / merged["monthly_budget"],
        0.0,
    )
    return merged


def build_cash_settings(
    *,
    company_name: str,
    base_currency: str,
    opening_cash: float,
    reserve_ratio: float,
    bank_accounts: list[str],
) -> pd.DataFrame:
    reserve_target = opening_cash * reserve_ratio
    return pd.DataFrame(
        [
            {"setting": "Company", "value": company_name},
            {"setting": "Base currency", "value": base_currency},
            {"setting": "Opening cash", "value": opening_cash},
            {"setting": "Liquidity reserve ratio", "value": reserve_ratio},
            {"setting": "Reserve target", "value": reserve_target},
            {"setting": "Bank / cash accounts", "value": ", ".join(bank_accounts)},
        ]
    )


def build_source_manifest(rows: list[dict[str, object]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(
            columns=["file_name", "file_type", "size_kb", "rows_detected", "status", "ocr_preview"]
        )
    return pd.DataFrame(rows)


def build_mapping_catalog() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "keywords": ", ".join(rule["keywords"]),
                "inferred_account": rule["account"],
                "budget_bucket": rule["bucket"],
                "cash_flow_class": rule["cash_flow"],
                "ifrs_hint": rule["ifrs"],
                "confidence": rule["confidence"],
            }
            for rule in ACCOUNT_RULES
        ]
    )


def build_accounting_summary(journal_df: pd.DataFrame) -> dict[str, float]:
    if journal_df is None or journal_df.empty:
        return {
            "entries": 0.0,
            "net_cash": 0.0,
            "revenue": 0.0,
            "expenses": 0.0,
        }

    revenue = journal_df.loc[journal_df["bucket_type"].eq("revenue"), "amount"].clip(lower=0.0).sum()
    expenses = journal_df.loc[
        journal_df["bucket_type"].isin(["expense", "capex", "financing"]),
        "amount",
    ].abs().sum()
    return {
        "entries": float(len(journal_df)),
        "net_cash": float(journal_df["cash_movement"].sum()),
        "revenue": float(revenue),
        "expenses": float(expenses),
    }


def build_accounting_workbook(
    *,
    journal_df: pd.DataFrame,
    budget_df: pd.DataFrame,
    budget_vs_actual_df: pd.DataFrame,
    cash_settings_df: pd.DataFrame,
    source_manifest_df: pd.DataFrame,
    company_name: str,
) -> bytes:
    overview = pd.DataFrame(
        [
            {"metric": "Company", "value": company_name},
            {"metric": "Entries", "value": len(journal_df)},
            {"metric": "Net cash movement", "value": journal_df["cash_movement"].sum() if not journal_df.empty else 0.0},
            {"metric": "Mapped ratio", "value": float((journal_df["mapping_confidence"] >= 0.8).mean()) if not journal_df.empty else 0.0},
        ]
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        journal_df.to_excel(writer, sheet_name="Normalized_Journal", index=False)
        budget_df.to_excel(writer, sheet_name="Budget_Template", index=False)
        budget_vs_actual_df.to_excel(writer, sheet_name="Budget_vs_Actual", index=False)
        cash_settings_df.to_excel(writer, sheet_name="Cash_Settings", index=False)
        source_manifest_df.to_excel(writer, sheet_name="Source_Files", index=False)
        build_mapping_catalog().to_excel(writer, sheet_name="Account_Mapping", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="163A5F")
        header_font = Font(color="FFFFFF", bold=True)

        for worksheet in workbook.worksheets:
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 36)

    output.seek(0)
    return output.getvalue()
